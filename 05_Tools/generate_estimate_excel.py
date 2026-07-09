#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
bravesoft 標準見積もり Excel 生成ツール

estimate_master.md（正本）の内容を JSON にした入力ファイルから、
社内標準の見積もりワークブック（.xlsx）を生成する。

シート構成（社内の実案件見積もり Excel を標準化したもの）:
  使い方          … 非エンジニア向けの操作ガイド（色の凡例・チェック列の説明）
  設定            … 単価表・管理費率・バッファ係数・税率・値引き・更新履歴（青セル＝入力欄）
  前提条件        … 見積根拠・対応範囲・対象外・契約形態・納品物（Rule 3。顧客提示の正本・独立シート）
  案比較          … 複数案（patterns が 2 つ以上）のときだけ生成
  見積サマリ[_案] … 案件概要＋金額サマリ（工数/金額税抜/税込/採算）。見積書の隣・数式連携。
                     開いた瞬間に合計が見える（下スクロール不要）。顧客提出は印刷範囲 A:D
  見積書[_案]     … 顧客に渡す見積明細（計上✓で金額に反映）。社内列（社内金額👤/根拠/MWO）と
                     右側の採算・抜け漏れ・マトリクスは折りたたみグループ（社内用）。
                     顧客提出は印刷範囲 A:I（社内列を除く）を PDF、または HTML/PDF 出力
  スケジュール[_案]… 開始日・並走人数・開始ラグで連動する動的ガント（条件付き書式）
  ヒアリングシート … 観点チェック（51+3 項目）。AI が既知の回答を事前記入・人は「確認」に✓（✓で行が色付く）
  非機能要件チェック … A〜H・29 観点 × 対象分野 × 判定 × 明細への反映先（Rule 2）
  機能×API対応表  … ハイブリッド/API 連携の特殊案件専用（JSON の feature_api 指定時のみ）

使い方:
  python3 generate_estimate_excel.py --input sample_input.json --out 見積もり.xlsx

入力 JSON の形（sample_input.json 参照）:
{
  "meta":   {"project","client","date","valid_until","precision","version"},
  "params": {"pm_rate":0.2,"buffer":1.2,"apply_buffer":false,"tax":0.1,
             "days_per_month":20,"discount":0},
  "rates":  [{"key":"PM","name":"プロジェクトマネージャー",
              "internal":60000,"external":75000}, ...],
  "patterns":[{"name":"Phase1",            # 1 案なら name 省略可
               "items":[
                 {"type":"pm","phase":"要件定義","item":"要件定義管理",
                  "desc":"...","rate":null},          # rate 省略時は 設定の管理費率
                 {"phase":"要件定義","category":"要件定義","item":"プロジェクト計画書",
                  "desc":"...","days":3,"role":"PM","impl":"－",
                  "basis":"...","mwo":"M","checked":true},
                 {"type":"parent","phase":"開発","category":"アプリ","item":"ログイン機能",
                  "desc":"...","children":[ {leaf...}, ... ]}
               ],
               "schedule":[{"task","owner","role","days","lag","biz_days"}]  # 省略で自動生成
             }],
  "feature_api":[{"area","detail","apis","api_count","p1","p2","full",
                  "impl","note"}, ...]                                  # 任意
}

設計メモ（数式の頑健性）:
- 明細の BH 列（非表示）に行種別 leaf/pm/parent を持ち、全集計は BH 列の完全一致
  SUMIFS（全列参照）で行う → 行の追加/挿入に追随し、項目名の文字列に依存しない。
- BI/BJ 列（非表示）= leaf 行のみの社内/社外金額。PM 行はこの列を SUMIFS するため
  自己参照（循環）せずに「フェーズ費用×率」を全列参照で計算できる。
- 補助列を BH〜BJ に置くのは、右側の集計・フェーズ×カテゴリマトリクスが
  カテゴリ数だけ横に伸びても衝突しないため（O 列起点＋最大数十列を確保）。
- 担当が空欄/単価表に無い場合は金額 0＋担当セルを赤表示（#N/A を合計に伝播させない）。
- 金額は全て Excel 数式（単価変更・チェック ON/OFF で全体が連動する）。
- 社内金額・利益率を含むため、このワークブックは社内用（Rule 10）。
"""

import argparse
import copy
import json
import sys
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

# ---------------------------------------------------------------- ブランド定数
FONT = "Noto Sans JP"
C_PRIMARY = "0066FF"   # bravesoft key color
C_NAVY = "253A58"      # stats / dark surface
C_INK = "090A0A"
C_MUTED = "6C7072"
C_HAIRLINE = "E3E5E5"
C_SOFT = "E5F0FF"      # tonal blue = 入力欄
C_SURFACE = "F7F9FA"
C_CHIP = "F2F4F5"
C_SKY = "3F9CFF"
C_SUCCESS = "00A86B"
C_ERROR = "D63637"
C_ERROR_BG = "FBE9E9"
C_WHITE = "FFFFFF"

YEN = '[$¥-411]#,##0'
PCT = '0.0%'

F_BASE = Font(name=FONT, size=10, color=C_INK)
F_BOLD = Font(name=FONT, size=10, bold=True, color=C_INK)
F_HEAD = Font(name=FONT, size=10, bold=True, color=C_WHITE)
F_MUTED = Font(name=FONT, size=9, color=C_MUTED)
F_TITLE = Font(name=FONT, size=14, bold=True, color=C_INK)
F_SECTION = Font(name=FONT, size=11, bold=True, color=C_WHITE)
F_WARN = Font(name=FONT, size=10, bold=True, color=C_ERROR)

FILL_HEAD = PatternFill("solid", start_color=C_PRIMARY)
FILL_NAVY = PatternFill("solid", start_color=C_NAVY)
FILL_INPUT = PatternFill("solid", start_color=C_SOFT)
FILL_SURFACE = PatternFill("solid", start_color=C_SURFACE)
FILL_CHIP = PatternFill("solid", start_color=C_CHIP)

THIN = Side(style="thin", color=C_HAIRLINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

UNLOCKED = Protection(locked=False)

CHECK_ON = "✓"
CHECK_OFF = "－"
PARENT_MARK = "【内訳合計】"
PM_CATEGORY = "PM・全体管理"

NOT_OFF = f'"<>{CHECK_OFF}"'
SLACK = 30  # 明細下部に確保する追記用の余白行（DV・条件付き書式・種別を先敷き）

# Rule 2 のカテゴリ軸（明細カテゴリ列のプルダウン候補。自由入力も許可）
CATEGORIES = ("PM・全体管理,要件定義,UI/UXデザイン,外部設計,内部設計,"
              "アプリ(iOS/Android),フロントエンド(Web),API・サーバ,管理画面(CMS),"
              "インフラ・環境,バッチ,外部連携,テスト・検証,セキュリティ,運用・保守,共通")

NFR_SHEET = "非機能要件チェック"
# 非機能要件チェック（nonfunctional_requirements_checklist.md と同じ A〜H・29 観点）
# (コード, 区分, 観点, 対象分野の目安)
NFR_DEFAULTS = [
    ("A1", "A. 性能・効率性", "画面表示・API レスポンス目標（例: 主要画面 2 秒以内）", "共通"),
    ("A2", "A. 性能・効率性", "同時接続・ピーク負荷（想定ユーザー数・イベント時の集中）", "サーバ・インフラ"),
    ("A3", "A. 性能・効率性", "データ量・件数の上限（一覧の件数・画像/動画サイズ）", "API・サーバ"),
    ("A4", "A. 性能・効率性", "負荷テスト・チューニングの要否", "サーバ・インフラ"),
    ("B1", "B. 可用性・信頼性", "稼働率（SLA）・メンテナンス時間", "インフラ"),
    ("B2", "B. 可用性・信頼性", "障害時の復旧（バックアップ・リストア・冗長化）", "インフラ"),
    ("B3", "B. 可用性・信頼性", "データ整合性（決済・在庫など金銭/重要データ）", "API・サーバ"),
    ("C1", "C. セキュリティ", "認証・認可（ログイン方式・権限・多要素・SSO）", "共通"),
    ("C2", "C. セキュリティ", "個人情報・機微情報の取扱い（暗号化・マスキング）", "API・サーバ"),
    ("C3", "C. セキュリティ", "決済・金銭処理のセキュリティ（PCI DSS 等）", "API・サーバ"),
    ("C4", "C. セキュリティ", "脆弱性診断・対応（リリース前診断の要否）", "共通"),
    ("C5", "C. セキュリティ", "通信の暗号化・証明書・WAF", "インフラ"),
    ("D1", "D. 運用・保守性", "ログ・監視・アラート（障害検知・通知）", "インフラ"),
    ("D2", "D. 運用・保守性", "運用マニュアル・管理者向け操作手順", "共通"),
    ("D3", "D. 運用・保守性", "保守・問い合わせ対応（保証期間・SLA・別契約）", "共通"),
    ("D4", "D. 運用・保守性", "リリース・デプロイ運用（手順書・ロールバック）", "共通"),
    ("E1", "E. 拡張性・移植性", "将来の機能拡張・スケール（マルチテナント等）", "API・サーバ"),
    ("E2", "E. 拡張性・移植性", "対応 OS・端末・ブラウザの範囲（世代・機種）", "アプリ/Web"),
    ("E3", "E. 拡張性・移植性", "データ移行（既存システムからの移行）", "API・サーバ"),
    ("E4", "E. 拡張性・移植性", "多言語・多通貨・地域対応", "共通"),
    ("F1", "F. 互換性・外部連携", "外部 API・他社サービス連携（決済・地図・SNS・LINE 等）", "外部連携"),
    ("F2", "F. 互換性・外部連携", "既存システム・基盤との連携（ID基盤・CMS・在庫）", "外部連携"),
    ("F3", "F. 互換性・外部連携", "データ入出力（CSV/帳票/分析基盤・BI）", "API・サーバ"),
    ("G1", "G. 法令・規約・ストア", "App Store / Google Play 審査要件・掲載素材", "アプリ"),
    ("G2", "G. 法令・規約・ストア", "利用規約・プライバシーポリシー・同意取得", "共通"),
    ("G3", "G. 法令・規約・ストア", "業界・法令対応（特商法・景表法・資金決済法 等）", "共通"),
    ("G4", "G. 法令・規約・ストア", "アクセシビリティ要件", "アプリ/Web"),
    ("H1", "H. 分析・計測", "アクセス解析・行動ログ（GA・独自イベント）", "共通"),
    ("H2", "H. 分析・計測", "KPI ダッシュボード・レポート", "管理画面"),
]


# ---------------------------------------------------------------- helpers
def st(cell, value=None, font=F_BASE, fill=None, fmt=None, align=None,
       border=True, wrap=False, unlock=False):
    """set value + style in one call"""
    if value is not None:
        cell.value = value
    cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    cell.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        cell.border = BORDER
    if unlock:
        cell.protection = UNLOCKED
    return cell


def protect(ws):
    """シート保護（パスワード無し）。行の挿入/削除・書式変更は許可する。"""
    ws.protection.sheet = True
    ws.protection.insertRows = False   # False = 許可
    ws.protection.deleteRows = False
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False


def sheet_suffix(pattern_name):
    return f"_{pattern_name}" if pattern_name else ""


def parse_date(s):
    if isinstance(s, (date, datetime)):
        return s
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return s  # 文字列のまま置く（自由記述を許容）


def flatten_items(items):
    """children 入れ子を平坦化し、行種別(kind)と親子関係を確定する。
    各フェーズ内で pm 行は先頭に並べ替える。"""
    flat = []
    for it in items:
        kind = it.get("type", "normal")
        if kind == "parent":
            children = it.get("children", [])
            if not children:
                raise ValueError(f"親行に children がありません: {it.get('item')}")
            parent = {k: v for k, v in it.items() if k != "children"}
            parent["kind"] = "parent"
            parent["n_children"] = len(children)
            flat.append(parent)
            for ch in children:
                c = dict(ch)
                c["kind"] = "leaf"
                c["child"] = True
                c.setdefault("phase", it.get("phase"))
                c.setdefault("category", it.get("category"))
                flat.append(c)
        elif kind == "pm":
            p = dict(it)
            p["kind"] = "pm"
            p.setdefault("category", PM_CATEGORY)
            flat.append(p)
        else:
            l = dict(it)
            l["kind"] = "leaf"
            flat.append(l)

    # 種別は S 列で判定するため、項目名にマーカー文字列が紛れるのは禁止
    for it in flat:
        if it["kind"] != "parent" and PARENT_MARK in str(it.get("item", "")):
            raise ValueError(
                f"通常行の項目名に {PARENT_MARK} を含めない: {it['item']}")

    # フェーズ出現順を保ちつつ、フェーズ内で pm を先頭へ
    phases = []
    for it in flat:
        if it["phase"] not in phases:
            phases.append(it["phase"])
    ordered = []
    for ph in phases:
        in_phase = [it for it in flat if it["phase"] == ph]
        pms = [it for it in in_phase if it["kind"] == "pm"]
        rest = [it for it in in_phase if it["kind"] != "pm"]
        ordered.extend(pms + rest)
    return ordered, phases


# ---------------------------------------------------------------- 使い方シート
def build_readme(wb, has_multi_pattern, has_feature_api=False):
    ws = wb.create_sheet("使い方")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 96

    r = 2
    st(ws.cell(r, 2), "📖 この見積もりファイルの使い方", F_TITLE, border=False)
    r += 1
    st(ws.cell(r, 2), "bravesoft 標準見積もり — 誰でも 3 ステップで金額を動かせます",
       F_MUTED, border=False)
    r += 2

    steps = [
        ("1️⃣ 設定", "「⚙ 設定」シートの青いセルを埋める（単価・管理費率・バッファ・税率・値引き）。ここを変えると全シートの金額が自動で変わります。"),
        ("2️⃣ 前提条件", "「📋 前提条件」シートを埋める（見積根拠・対応範囲・対象外・契約形態・納品物）。金額の前提なので必須。顧客説明ではこのシートを先に提示します（独立シート）。"),
        ("3️⃣ 明細と計上✓", "「見積書」シートの見積明細で工数・担当・項目を調整。金額から外したい行は A 列「計上」を ✓ → － に（プルダウン）。その行はグレーになり、隣の「見積サマリ」・PM費・税込合計・スケジュールまで全部が自動で再計算されます。行を増やすときは既存の明細行をコピーして途中に挿入（一番下の行の下に足さない）。"),
    ]
    for label, text in steps:
        st(ws.cell(r, 2), label, Font(name=FONT, size=11, bold=True, color=C_PRIMARY), border=False)
        st(ws.cell(r, 3), text, F_BASE, border=False, wrap=True)
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1

    # ⏭ 全体の流れ（キットの Rule 12: 常に「次に何をするか」を示す）
    st(ws.cell(r, 2), "⏭ 全体の流れ", F_BOLD, border=False)
    st(ws.cell(r, 3),
       "① ⚙ 設定を埋める → ② 📋 前提条件・対象外を埋める → ③ 「見積書」で明細を調整（計上 ✓/－ でスコープ・金額を調整）→ ④ 「見積サマリ」で金額を確認 → "
       "④ 🛡 非機能要件チェックと 🔍 抜け漏れチェックを全部 ✅ に → "
       "⑤ 👤 集計・採算で利益率を確認 → ⑥ 🔔 上長・PM の人間レビュー（Rule 13・必須）→ "
       "⑦ 📤 顧客提出は AI に「提出用の HTML / PDF を生成して」と依頼、または「見積書」の社内列を折りたたんで印刷範囲 A:I を PDF（この Excel は社内用）",
       F_BASE, border=False, wrap=True)
    ws.row_dimensions[r].height = 44
    r += 1
    st(ws.cell(r, 3),
       "迷ったら AI に聞けば OK。キットは 🧭(開始) 📋(作業開始) ❓(質問) 🧮(試算) 🔔(要レビュー) "
       "✅⏭(完了と次アクション) ⚠️(ブロッカー) 📤(出力) の絵文字で、毎回「次に何をすればいいか」を案内します。",
       F_MUTED, border=False, wrap=True)
    ws.row_dimensions[r].height = 30
    r += 2

    st(ws.cell(r, 2), "🎨 色の凡例", F_BOLD, border=False)
    r += 1
    legends = [
        (FILL_INPUT, None, "うすい水色のセル＝入力欄（自由に変更してよい）"),
        (None, None, "白いセル＝自動計算や文字の欄。金額の白セルは保護済み（さわらない）。項目名・概要・根拠・担当・計上などの文字欄は白でも編集OK"),
        (FILL_CHIP, None, "グレー帯＝小計・集計行"),
        (None, Font(name=FONT, size=10, color="9AA0A6", strike=True), "グレー文字＋取り消し線の行＝「計上 －」で金額から除外中の行"),
        (FILL_NAVY, None, "紺色＝見出し・強調（入力欄ではない）"),
        (PatternFill("solid", start_color=C_ERROR_BG), None, "赤いセル＝担当が未入力か、設定の単価表に無い役割（金額が 0 になっています。担当を直してください）"),
    ]
    for fill, font, text in legends:
        c = ws.cell(r, 2)
        if fill:
            c.fill = fill
        if font:
            st(c, "サンプル", font, fill, border=True)
        c.border = BORDER
        st(ws.cell(r, 3), text, F_BASE, border=False, wrap=True)
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1

    notes = [
        ("📊 「見積サマリ」シート", "見積書の隣にある、案件概要＋金額サマリ（フェーズ別の工数・金額税抜、合計税抜/消費税/税込、採算）のシート。見積書と数式連携しているので、明細を直すと自動で更新されます。開いた瞬間に合計が見えるので、下にスクロールする必要はありません。顧客提出は印刷範囲 A:D（採算・社内数字は範囲外）。"),
        ("📝 「見積書」シート", "顧客に渡す見積明細。工数・担当・項目を調整し、計上 ✓/－ で金額に反映します。編集はここ、金額の確認は隣の「見積サマリ」で。"),
        ("📕 社内列の折りたたみ", "見積書の 社内金額 👤・根拠・優先度(MWO) の列（J〜L）と右側の分析ブロックは『社内グループ』にまとまっています。列の上の [－] ボタンでワンクリックで隠せます。顧客に見せる列（項目〜金額税抜＝A〜I）だけが残るので、印刷範囲 A:I をそのまま PDF にすれば顧客提出物になります（社内の原価・利益率は出ません）。見積サマリも同様に A:D が顧客提出範囲です。"),
        ("✓ チェック列（計上）", "✓＝金額に入れる ／ －＝外す。スコープ調整・値引き交渉のときは、行を消さずに － にするのがおすすめ（「やらないこと」が残るので後で戻せる・説明できる）。★スコープ未確定の機能も － のまま一覧に残せます。"),
        ("📦 【内訳合計】の親行", "大きな機能を子行（└）に分解したときの親行。金額・工数は子行の合計が自動で入ります。機能ごと外したいときは、親行ではなく子行（└）すべての計上を － にしてください（親行の合計が自動で 0 になります）。子行を増やすときは子行と子行の間に挿入。"),
        ("🧮 PM・全体管理の行", "工数欄が % になっている行は「そのフェーズの費用 × 率」で自動計算される管理費です。率は設定シートの管理費率と連動。行ごとに数字で上書きもできます（設定連動に戻すには、そのセルに半角で =管理費率 と入力）。"),
        ("🔤 優先度（MWO）列", "金額を動かすいちばん大事な列。M(Must)=必須＝これだけで成立する最小構成 ／ W(Want)=できれば欲しい＝【予算調整の余地】 ／ O(Out)=対象外 ／ ★未確定=要件が固まっていない（0 計上で一覧に残す）。明細右の 💡 ブロックに「M だけの金額」「W の合計（＝いくら削れるか）」が自動集計されるので、予算オーバーのときは ① 💡 で W の合計を見る → ② W の行の計上を － にする、の順で調整します（値引きより先に）。W は青太字・★未確定は赤で表示されます。"),
        ("💰 バッファ・体制維持費・値引き", "設定シートで「提示金額にバッファを適用」を ✓ にすると、提示額に係数分が上乗せされます。体制維持費（任意・率）と値引きも設定シートの値が提示額に反映されます。全部「見える行」として金額構成に出ます（Rule 5: 盛りと儲けを隠さない）。"),
        ("📚 ことば（これだけ）", "人日＝1 人が 1 営業日（約 8 時間）働く量。3 人日＝1 人なら 3 日。／ 社内単価＝原価（社外秘）・社外単価＝お客様への請求単価。提示額に効くのは社外単価。／ 利益率 2 種＝上段は予定どおり終わった場合、下段はバッファ分まで工数を使い切った場合の最悪ライン。下段がマイナスなら受注前に要相談。"),
        ("👤 社内限りマーク", "社内金額・利益率・社内単価は社外に出せません（Rule 10）。このファイルは【社内用】です。顧客提出はキットの提出用 HTML / PDF が標準。Excel で渡す場合は「見積書」の社内列（J〜L と右側ブロック）を折りたたみ、印刷範囲 A:I を PDF 出力してください（社内数字は出ません）。"),
        ("📅 スケジュール（動的）", "開始日・並走人数・開始ラグ・審査待ち日数（青セル）を変えると、ガントと総期間が自動で動きます。「工数(人日)」が数式の工程は「見積書」の「見積書」の見積明細と連動しており、明細の工数変更や計上 － でスケジュールも自動で伸縮します。開始ラグを手入力すると、その工程以降の「前の工程が終わったら開始」の自動連結が外れます（並走させたいときだけ上書き。戻すには =前の行のG+前の行のF）。ガントの月列は 18 ヶ月分。超える場合は右に列をコピーしてください。"),
        ("📋 前提条件・対象外", "見積根拠・対応範囲・対象外・契約形態・環境・納品物・検収条件を記録する独立シート（Rule 3）。「ここから外れたら別途お見積もり」の線引きで、あとから揉めないための保険。顧客説明ではまずこのシートを提示し、その後「見積書」で金額の内訳を説明する運びが基本。"),
        ("🗣 ヒアリングシート", "お客様ヒアリングの観点漏れを防ぐチェック（51+3 項目）。見積もりから分かる項目（例: オフショア利用・管理画面・多言語・脆弱性診断 等）は AI が回答を事前記入（水色セル＝要確認）。全項目に答える必要はなく、人は各観点を見て『確認』列に ✓ を付けるだけ（✓ で行が緑に色付き、上部に「確認済み ○/54」が出ます）。"),
        ("🛡 非機能要件チェック", "性能・可用性・セキュリティ・運用・拡張性・外部連携・法令/ストア・分析の 29 観点を、対象分野（アプリ/Web/API・サーバ/インフラ等）ごとに 対象/対象外/要確認 で判定するシート。「対象」にしたら見積明細に工数を積んで反映先を記入（未反映は赤くなる）。見積明細の 🔍 抜け漏れチェックと連動しており、未判定が残ると ⚠ が出ます。"),
        ("🔍 抜け漏れチェック", "「見積書」右側（社内グループ内）で、PM・テスト・リリース/ストア申請・非機能の計上を自動判定（Rule 2）。その下の「フェーズ×カテゴリ 工数マトリクス」で、空欄（＝工数ゼロ）の交点に抜けがないかを一目で確認できます。カテゴリ列はプルダウンから選択（自由入力も可）。"),
        ("🔒 シート保護について", "数式セルは誤上書き防止のため保護しています（パスワード無し）。どうしても直したいときは「校閲 → シート保護の解除」。"),
        ("♻️ 作り直したいとき", "このファイルは正本（estimate_master.md）から AI が生成しています。行の大量追加・フェーズ追加は、AI に「見積もり Excel を再生成して」と依頼するのが確実です。"),
    ]
    if has_multi_pattern:
        notes.append(("🗂 複数案", "見積書_○○ のように案（Phase 別・スコープ別）ごとにシートが分かれています。設定・前提条件シートは全案で共通。「案比較」シートで工数・金額・利益率 👤 を横並びで比較できます。"))
    if has_feature_api:
        notes.append(("🔗 機能×API対応表", "この案件専用の任意シート（ハイブリッド / API 連携案件向け）。機能↔API↔実装方式↔Phase の漏れ逆引き用で、工数の正は見積明細です。"))
    for label, text in notes:
        st(ws.cell(r, 2), label, F_BOLD, border=False, wrap=True)
        st(ws.cell(r, 3), text, F_BASE, border=False, wrap=True)
        ws.row_dimensions[r].height = max(30, 15 * (len(text) // 60 + 1))
        r += 1
    return ws


# ---------------------------------------------------------------- 設定シート
def build_settings(wb, meta, params, rates):
    ws = wb.create_sheet("設定")
    ws.sheet_view.showGridLines = False
    widths = {"A": 30, "B": 18, "C": 14, "D": 14, "E": 14, "F": 12, "G": 52}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    st(ws.cell(1, 1), "⚙ 見積もり設定（青いセルだけ変更すれば OK）", F_TITLE, border=False)

    st(ws.cell(3, 1), "■ 案件情報", F_BOLD, border=False)
    valid_until = meta.get("valid_until")
    meta_rows = [
        ("案件名", meta.get("project", ""), None),
        ("顧客名（宛先）", meta.get("client", ""), None),
        ("作成日", parse_date(meta.get("date", "")), "yyyy/mm/dd"),
        ("見積有効期限", parse_date(valid_until) if valid_until else "=$B$6+30",
         "yyyy/mm/dd"),
        ("見積精度", meta.get("precision", "概算 −20%〜+30%"), None),
        ("バージョン", meta.get("version", "1.0"), None),
    ]
    r = 4
    for label, value, fmt in meta_rows:
        st(ws.cell(r, 1), label, F_BASE)
        st(ws.cell(r, 2), value, F_BASE, FILL_INPUT, fmt=fmt, unlock=True)
        r += 1
    st(ws.cell(10, 1), "作成者 👤（社内用・提出物には出さない）", F_BASE)
    st(ws.cell(10, 2), meta.get("author", ""), F_BASE, FILL_INPUT, unlock=True)

    st(ws.cell(11, 1), "■ 係数・率（Rule 5: 盛りと儲けを見える化）", F_BOLD, border=False)
    param_rows = [  # (row, label, value, fmt, note)
        (12, "管理費率（PM費）", params.get("pm_rate", 0.2), PCT,
         "フェーズ費用に掛ける率。標準 20〜30%（pricing_and_rates.md）"),
        (13, "バッファ係数", params.get("buffer", 1.2), "0.0",
         "予備工数の係数。標準 1.2x"),
        (14, "提示金額にバッファを適用", CHECK_ON if params.get("apply_buffer") else CHECK_OFF, None,
         "✓ にすると提示額（税抜）に係数分を上乗せ。－ なら精度レンジで吸収"),
        (15, "消費税率", params.get("tax", 0.10), "0%", "標準 10%"),
        (16, "営業日/月", params.get("days_per_month", 20), "0", "人月換算・スケジュールで使用"),
        (17, "値引き（税抜・正の数で入力）", params.get("discount", 0), YEN,
         "第一候補は値引きでなくスコープ調整（計上チェックを － に）"),
    ]
    param_rows.append(
        (18, "体制維持費率（任意）", params.get("retainer_rate", 0), PCT,
         "体制を確保して請ける場合に小計へ上乗せする率（例 5%）。不要なら 0%"))
    for r, label, value, fmt, note in param_rows:
        st(ws.cell(r, 1), label, F_BASE)
        st(ws.cell(r, 2), value, F_BASE, FILL_INPUT, fmt=fmt, unlock=True)
        st(ws.cell(r, 7), note, F_MUTED, border=False)
    dv_check = DataValidation(type="list", formula1=f'"{CHECK_ON},{CHECK_OFF}"',
                              allow_blank=True)
    ws.add_data_validation(dv_check)
    dv_check.add(ws.cell(14, 2))
    dv_disc = DataValidation(
        type="decimal", operator="greaterThanOrEqual", formula1="0",
        allow_blank=True, showErrorMessage=True,
        errorTitle="値引きの入力", error="値引きは正の数で入力してください（例: 100000）")
    ws.add_data_validation(dv_disc)
    dv_disc.add(ws.cell(17, 2))

    st(ws.cell(19, 1), "■ 単価表（役割別・円/人日）", F_BOLD, border=False)
    st(ws.cell(19, 7),
       "社内単価＝原価（👤社内限り・Rule 10）。⚠ 初期値は暫定 — 案件・経理の基準で必ず見直す",
       F_MUTED, border=False)
    headers = ["役割キー", "役割名", "社内単価 👤", "社外単価", "人月換算(社外)", "粗利率 👤"]
    for i, h in enumerate(headers, start=1):
        st(ws.cell(20, i), h, F_HEAD, FILL_HEAD, align="center")
    st(ws.cell(20, 7),
       "⚠ 役割キー（A列）は見積明細の「担当」と連動。既存キーは変更しない（表示名を変えたいときは B 列を編集）",
       F_MUTED, border=False)
    RATE_TOP, RATE_ROWS = 21, 10
    for i in range(RATE_ROWS):
        r = RATE_TOP + i
        rate = rates[i] if i < len(rates) else None
        st(ws.cell(r, 1), rate["key"] if rate else None, F_BASE, FILL_INPUT, unlock=True)
        st(ws.cell(r, 2), rate.get("name", "") if rate else None, F_BASE, FILL_INPUT, unlock=True)
        st(ws.cell(r, 3), rate["internal"] if rate else None, F_BASE, FILL_INPUT, YEN, unlock=True)
        st(ws.cell(r, 4), rate["external"] if rate else None, F_BASE, FILL_INPUT, YEN, unlock=True)
        st(ws.cell(r, 5),
           f'=IF($D{r}="","",$D{r}*営業日月)', F_BASE, fmt=YEN)
        st(ws.cell(r, 6),
           f'=IF(OR($D{r}="",$D{r}=0),"",($D{r}-$C{r})/$D{r})', F_BASE, fmt=PCT)
    rate_end = RATE_TOP + RATE_ROWS - 1

    st(ws.cell(rate_end + 2, 1),
       "※ 行が足りなければ空行に役割を追記（見積明細の担当プルダウンに自動反映）。", F_MUTED, border=False)

    st(ws.cell(rate_end + 4, 1), "■ 更新履歴（版管理: 金額や前提を変えたら 1 行残す）",
       F_BOLD, border=False)
    hist_headers = ["日付", "版", "担当", "変更内容"]
    for i, h in enumerate(hist_headers, start=1):
        st(ws.cell(rate_end + 5, i), h, F_HEAD, FILL_HEAD, align="center")
    meta_hist_top = rate_end + 6
    st(ws.cell(meta_hist_top, 1), parse_date(meta.get("date", "")), F_BASE, FILL_INPUT,
       "yyyy/mm/dd", unlock=True)
    st(ws.cell(meta_hist_top, 2), meta.get("version", "1.0"), F_BASE, FILL_INPUT, unlock=True)
    st(ws.cell(meta_hist_top, 3), meta.get("author", ""), F_BASE, FILL_INPUT, unlock=True)
    st(ws.cell(meta_hist_top, 4), "初版作成（AI 生成）", F_BASE, FILL_INPUT, unlock=True)
    for rr in range(meta_hist_top + 1, meta_hist_top + 4):
        for c in range(1, 5):
            st(ws.cell(rr, c), "", F_BASE, FILL_INPUT, unlock=True)

    st(ws.cell(meta_hist_top + 5, 1),
       "⏭ 次にやること: 設定が埋まったら「前提条件」→「見積書」シートへ（青セル以外は自動計算）",
       Font(name=FONT, size=10, bold=True, color=C_PRIMARY), border=False)

    # 定義名（全シートの数式から日本語名で参照する）
    defs = {
        "管理費率": "設定!$B$12",
        "バッファ係数": "設定!$B$13",
        "バッファ適用": "設定!$B$14",
        "税率": "設定!$B$15",
        "営業日月": "設定!$B$16",
        "値引き": "設定!$B$17",
        "体制維持費率": "設定!$B$18",
        "単価表": f"設定!$A${RATE_TOP}:$F${rate_end}",
        "役割一覧": f"設定!$A${RATE_TOP}:$A${rate_end}",
    }
    for name, ref in defs.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))
    protect(ws)
    return ws


# ---------------------------------------------------------------- 前提条件・スコープ
ASSUMPTION_SECTIONS = [  # (JSON キー, 見出し, list か str か, 既定値)
    ("basis", "見積根拠（何を基に見積もったか）", True,
     ["（記入例）お客様ご提供の機能一覧・画面イメージ（YYYY-MM-DD 受領）"]),
    ("scope_in", "対応範囲（この見積もりに含む工程）", True,
     ["要件定義 ／ デザイン ／ 設計 ／ 開発 ／ テスト ／ リリース（PM・全体管理を含む）"]),
    ("scope_out", "対象外（Out of Scope — やらないこと）", True,
     ["（記入例）運用・保守（別途契約）", "（記入例）既存システム側の改修"]),
    ("contract", "契約形態", False, "請負（上流を準委任に分ける場合はここに記載）"),
    ("environment", "環境（対応 OS・端末・ブラウザ・言語）", True,
     ["（記入例）iOS / Android 各直近 2 世代・主要端末"]),
    ("deliverables", "納品物", True,
     ["（記入例）要件定義書・画面仕様書・テスト結果報告書・ビルド成果物"]),
    ("acceptance", "検収条件", False, "（記入例）納品後 10 営業日以内に検収"),
    ("source_code", "ソースコードの扱い", False, "（記入例）原則非納品（要協議）"),
    ("notes", "その他の前提", True,
     ["ここに示す前提から異なる場合は、別途お見積もりが必要となります。"]),
]


def build_assumptions(wb, assumptions):
    """前提条件・スコープシート（Rule 3: ここから外れたら別途）。
    顧客説明ではこのシートを先に提示する運用（Rule 3・独立シート）。"""
    ws = wb.create_sheet("前提条件")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 96

    st(ws.cell(1, 1), "📋 前提条件・スコープ（Rule 3: この金額の前提。外れたら別途お見積もり）",
       F_TITLE, border=False)
    st(ws.cell(2, 1),
       "青いセルを必ず埋める（AI に依頼して埋めても OK）。顧客説明ではまずこのシートを提示します。"
       "特に「対象外」が空だと、後から「それも含まれていると思った」トラブルになります。",
       F_MUTED, border=False)

    a = assumptions or {}
    refs = []   # (見出し, [(row)] )
    arow = {}   # key -> 先頭の値セル行（見積サマリの案件概要リンク用）
    r = 3
    for key, label, is_list, default in ASSUMPTION_SECTIONS:
        r += 1
        arow[key] = r
        val = a.get(key, default)
        items = val if is_list else [val]
        if not isinstance(items, list):
            items = [items]
        rows = []
        for i, item in enumerate(items):
            st(ws.cell(r, 1), ("■ " + label) if i == 0 else "", F_BOLD, wrap=True)
            st(ws.cell(r, 2), item, F_BASE, FILL_INPUT, wrap=True, unlock=True)
            ws.row_dimensions[r].height = 24
            rows.append(r)
            r += 1
        # 追記用に各セクション 1 行の空き（青・入力可）
        st(ws.cell(r, 1), "", F_BASE)
        st(ws.cell(r, 2), "", F_BASE, FILL_INPUT, wrap=True, unlock=True)
        rows.append(r)
        r += 1
        refs.append((label, rows))
        if key == "contract":
            dv = DataValidation(
                type="list",
                formula1='"請負,準委任,上流＝準委任 ／ 下流＝請負,併用（詳細は備考）"',
                allow_blank=True, showErrorMessage=False)
            ws.add_data_validation(dv)
            dv.add(ws.cell(rows[0], 2))

    st(ws.cell(r + 1, 1),
       "⏭ 次にやること: 全セクションを埋めたら「見積書」へ。対象外に迷ったら AI に"
       "「この案件の対象外の候補を挙げて」と依頼",
       Font(name=FONT, size=10, bold=True, color=C_PRIMARY), border=False)
    protect(ws)
    return arow


# ---------------------------------------------------------------- 見積明細シート
def build_detail(wb, pattern, phases, flat):
    """見積明細シート「見積書」（顧客に渡す明細＝提出物そのもの）。
    左＝顧客に見せる列（項目/概要/実装方式/工数/金額税抜=社外）を主役に、
    社内列（社内金額👤/根拠/優先度MWO）と右側の採算・分析ブロックは
    折りたたみグループ（社内用・ワンクリックで隠せる）。
    金額サマリ・案件概要は隣の「見積サマリ」シート（数式連携）に切り出す。"""
    name = "見積書" + sheet_suffix(pattern.get("name"))
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    widths = {"A": 6, "B": 11, "C": 14, "D": 38, "E": 50, "F": 11, "G": 9,
              "H": 12, "I": 14, "J": 14, "K": 30, "L": 12, "M": 2,
              "N": 24, "O": 12, "P": 13, "Q": 13, "R": 11,
              "BH": 8, "BI": 12, "BJ": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for col in ("BH", "BI", "BJ"):
        ws.column_dimensions[col].hidden = True

    pat = f"（{pattern['name']}）" if pattern.get("name") else ""
    # ---- 顧客向けヘッダ（表紙）
    st(ws.cell(1, 1), f'="御見積書｜"&設定!$B$4&"{pat}"', F_TITLE, border=False)
    st(ws.cell(2, 1), '=設定!$B$5&" 御中"', F_BOLD, border=False)
    st(ws.cell(3, 1),
       '="作成日: "&TEXT(設定!$B$6,"yyyy/mm/dd")&"　有効期限: "&TEXT(設定!$B$7,"yyyy/mm/dd")'
       '&"　見積精度: "&設定!$B$8&"　Ver."&設定!$B$9', F_MUTED, border=False)
    st(ws.cell(4, 1),
       "⚠ 社内用（J列＝社内金額・右側＝採算/利益率を含む）。このファイルのままお客様へ送らないこと。"
       "顧客提出はキットの HTML / PDF、または「社内列」を折りたたんで印刷範囲 A:I を PDF 出力（Rule 10）。",
       F_WARN, border=False)

    NF = NOT_OFF
    # ---- 見積明細（このシートの主役）。金額サマリ・案件概要は隣の「見積サマリ」シート（連携）へ
    st(ws.cell(6, 1), "📝 見積明細（工数×単価の積み上げ・計上✓で金額に反映）", F_BOLD, border=False)
    st(ws.cell(6, 10), "👤 社内金額は社内限り（Rule 10）／合計・サマリは隣の「見積サマリ」シート", F_MUTED, border=False)
    header_row = 7
    headers = ["計上", "フェーズ", "カテゴリ", "項目", "概要", "実装方式",
               "工数(人日)", "担当", "金額(税抜)", "社内金額 👤", "根拠・備考", "優先度(MWO)"]
    for i, h in enumerate(headers, start=1):
        st(ws.cell(header_row, i), h, F_HEAD, FILL_HEAD, align="center")
    st(ws.cell(header_row, 60), "行種別", F_MUTED, align="center")
    st(ws.cell(header_row, 61), "leaf社内", F_MUTED, align="center")
    st(ws.cell(header_row, 62), "leaf社外", F_MUTED, align="center")
    ws.cell(header_row, 1).comment = Comment(
        "✓＝金額に入れる／－＝金額から外す（プルダウン）。－にすると小計・PM費・税込合計・"
        "上部サマリ・スケジュールまで自動で再計算されます。", "estimation-kit")
    ws.cell(header_row, 9).comment = Comment(
        "金額(税抜)＝工数×担当の社外単価（お客様への請求ベース＝提示額のもと・Rule 1）。", "estimation-kit")
    ws.cell(header_row, 10).comment = Comment(
        "社内金額＝工数×担当の社内単価（原価・社内限り Rule 10）。この列は折りたたみグループ内。",
        "estimation-kit")
    ws.cell(header_row, 12).comment = Comment(
        "金額を動かすいちばん大事な列。M=必須／W=できれば（予算調整の余地）／O=対象外／"
        "★未確定=0計上で残す。予算相談ではまず W の行の計上を － に。", "estimation-kit")

    dv_check = DataValidation(type="list", formula1=f'"{CHECK_ON},{CHECK_OFF}"', allow_blank=True)
    dv_role = DataValidation(type="list", formula1="役割一覧", allow_blank=True)
    dv_impl = DataValidation(type="list",
                             formula1='"ネイティブ,WebView,外部ブラウザ,共通,－"', allow_blank=True)
    dv_mwo = DataValidation(type="list", formula1='"M,W,O,★未確定"', allow_blank=True)
    dv_cat = DataValidation(type="list", formula1=f'"{CATEGORIES}"',
                            allow_blank=True, showErrorMessage=False)
    for dv in (dv_check, dv_role, dv_impl, dv_mwo, dv_cat):
        ws.add_data_validation(dv)

    row_of = {}
    r = header_row
    for idx, it in enumerate(flat):
        r += 1
        row_of[idx] = r
        kind = it["kind"]
        fill = FILL_SURFACE if kind in ("pm", "parent") else None
        font = F_BOLD if kind == "parent" else F_BASE
        st(ws.cell(r, 60), kind, F_MUTED, align="center")
        # A 計上
        if kind == "parent":
            st(ws.cell(r, 1), "", font, fill, align="center")
        else:
            st(ws.cell(r, 1), CHECK_ON if it.get("checked", True) else CHECK_OFF,
               font, fill, align="center", unlock=True)
            dv_check.add(ws.cell(r, 1))
        # B フェーズ / C カテゴリ
        st(ws.cell(r, 2), it.get("phase", ""), font, fill, wrap=True, unlock=True)
        st(ws.cell(r, 3), it.get("category", ""), font, fill, wrap=True, unlock=True)
        if kind == "leaf":
            dv_cat.add(ws.cell(r, 3))
        # D 項目
        label = it.get("item", "")
        if kind == "parent" and PARENT_MARK not in label:
            label += PARENT_MARK
        if it.get("child") and not label.startswith("　└"):
            label = "　└ " + label
        st(ws.cell(r, 4), label, font, fill, wrap=True, unlock=True)
        # E 概要 / F 実装方式
        st(ws.cell(r, 5), it.get("desc", ""), F_MUTED if kind != "parent" else font,
           fill, wrap=True, unlock=True)
        st(ws.cell(r, 6), it.get("impl", "－"), font, fill, align="center", unlock=True)
        if kind == "leaf":
            dv_impl.add(ws.cell(r, 6))
        # G 工数 / H 担当 / I 金額(税抜=社外) / J 社内金額
        if kind == "pm":
            rate = it.get("rate")
            st(ws.cell(r, 7), rate if rate is not None else "=管理費率",
               F_BASE, FILL_INPUT, fmt="0%", align="right", unlock=True)
            st(ws.cell(r, 8), "－", F_BASE, fill, align="center")
            st(ws.cell(r, 9),
               f'=IF($A{r}="{CHECK_OFF}",0,ROUND(SUMIFS($BJ:$BJ,$B:$B,$B{r})*$G{r},0))',
               F_BASE, fill, YEN)
            st(ws.cell(r, 10),
               f'=IF($A{r}="{CHECK_OFF}",0,ROUND(SUMIFS($BI:$BI,$B:$B,$B{r})*$G{r},0))',
               F_BASE, fill, YEN)
        elif kind == "parent":
            n = it["n_children"]
            c1, c2 = r + 1, r + n
            st(ws.cell(r, 7), f'=SUMIFS(G{c1}:G{c2},A{c1}:A{c2},{NF})',
               font, fill, align="right")
            st(ws.cell(r, 8), "－", font, fill, align="center")
            st(ws.cell(r, 9), f"=SUM(I{c1}:I{c2})", font, fill, YEN)
            st(ws.cell(r, 10), f"=SUM(J{c1}:J{c2})", font, fill, YEN)
        else:
            st(ws.cell(r, 7), it.get("days", 0), F_BASE, FILL_INPUT, align="right", unlock=True)
            st(ws.cell(r, 8), it.get("role", ""), F_BASE, align="center", unlock=True)
            dv_role.add(ws.cell(r, 8))
            st(ws.cell(r, 9),
               f'=IF(OR($A{r}="{CHECK_OFF}",$H{r}=""),0,'
               f'IFERROR(ROUND($G{r}*VLOOKUP($H{r},単価表,4,FALSE),0),0))', F_BASE, fmt=YEN)
            st(ws.cell(r, 10),
               f'=IF(OR($A{r}="{CHECK_OFF}",$H{r}=""),0,'
               f'IFERROR(ROUND($G{r}*VLOOKUP($H{r},単価表,3,FALSE),0),0))', F_BASE, fmt=YEN)
            st(ws.cell(r, 61), f"=$J{r}", F_MUTED, fmt=YEN)   # BI = leaf 社内
            st(ws.cell(r, 62), f"=$I{r}", F_MUTED, fmt=YEN)   # BJ = leaf 社外
        # K 根拠 / L MWO
        st(ws.cell(r, 11), it.get("basis", ""), F_MUTED, fill, wrap=True, unlock=True)
        st(ws.cell(r, 12), it.get("mwo", "M" if kind == "leaf" else ""),
           font, fill, align="center", unlock=True)
        if kind == "leaf":
            dv_mwo.add(ws.cell(r, 12))

    last_row = r
    slack_end = last_row + SLACK
    for rr in range(last_row + 1, slack_end + 1):
        for col in (1, 2, 3, 4, 5, 6, 7, 8, 11, 12):
            ws.cell(rr, col).protection = UNLOCKED
    dv_check.add(f"A{last_row + 1}:A{slack_end}")
    dv_role.add(f"H{last_row + 1}:H{slack_end}")
    dv_impl.add(f"F{last_row + 1}:F{slack_end}")
    dv_mwo.add(f"L{last_row + 1}:L{slack_end}")
    dv_cat.add(f"C{last_row + 1}:C{slack_end}")

    # 条件付き書式（明細テーブル範囲のみ）
    grey = Font(name=FONT, size=10, color="9AA0A6", strike=True)
    hr1 = header_row + 1
    ws.conditional_formatting.add(
        f"A{hr1}:L{slack_end}",
        FormulaRule(formula=[f'$A{hr1}="{CHECK_OFF}"'], font=grey))
    ws.conditional_formatting.add(
        f"H{hr1}:H{slack_end}",
        FormulaRule(formula=[f'AND($BH{hr1}="leaf",$A{hr1}<>"{CHECK_OFF}",COUNTIF(役割一覧,$H{hr1})=0)'],
                    font=Font(name=FONT, size=10, bold=True, color=C_ERROR),
                    fill=PatternFill("solid", start_color=C_ERROR_BG)))
    for cond, f in (('="W"', Font(name=FONT, size=10, bold=True, color=C_PRIMARY)),
                    ('="O"', Font(name=FONT, size=10, color=C_MUTED)),
                    ('="★未確定"', Font(name=FONT, size=10, bold=True, color=C_ERROR))):
        ws.conditional_formatting.add(
            f"L{hr1}:L{slack_end}", FormulaRule(formula=[f'$L{hr1}{cond}'], font=f))

    # ---- 右側 社内ブロック（MWO別金額・抜け漏れ・マトリクス）＝ header_row に揃える
    rr = header_row
    st(ws.cell(rr, 14), "💡 優先度（MWO）別の金額 — 予算調整はここを見る 👤", F_BOLD, border=False)
    rr += 1
    for i, h in enumerate(["区分", "工数(計上分)", "", "社外金額", ""]):
        st(ws.cell(rr, 14 + i), h, F_HEAD, FILL_HEAD, align="center")
    def mwo_sum(col, code):
        return f'=SUMIFS(${col}:${col},$L:$L,"{code}",$BH:$BH,"leaf",$A:$A,{NF})'
    rr += 1
    st(ws.cell(rr, 14), "M（必須＝最小構成）", F_BOLD)
    st(ws.cell(rr, 15), mwo_sum("G", "M"), F_BASE, fmt="0.##", align="right")
    st(ws.cell(rr, 17), mwo_sum("I", "M"), F_BOLD, fmt=YEN)
    rr += 1
    st(ws.cell(rr, 14), "W（できれば＝調整余地）", Font(name=FONT, size=10, bold=True, color=C_PRIMARY))
    st(ws.cell(rr, 15), mwo_sum("G", "W"), F_BASE, fmt="0.##", align="right")
    st(ws.cell(rr, 17), mwo_sum("I", "W"),
       Font(name=FONT, size=10, bold=True, color=C_PRIMARY), fmt=YEN)
    rr += 1
    st(ws.cell(rr, 14), "O・★未確定（0円のはず）", F_BASE)
    ws.merge_cells(start_row=rr, start_column=14, end_row=rr, end_column=16)
    st(ws.cell(rr, 17),
       f'=IF(COUNTIFS($L:$L,"O",$A:$A,{NF},$I:$I,">0")'
       f'+COUNTIFS($L:$L,"★未確定",$A:$A,{NF},$I:$I,">0")=0,'
       f'"✅ 金額混入なし","⚠ 金額が入っている行あり")', F_BOLD, align="center")

    rr = header_row + 6
    st(ws.cell(rr, 14), "🔍 抜け漏れチェック（Rule 2・自動判定）👤", F_BOLD, border=False)
    rr += 1
    ok = '"✅ 計上あり"'
    checks = [
        ("PM・全体管理", f'=IF(SUMIFS($I:$I,$BH:$BH,"pm",$A:$A,{NF})>0,{ok},"⚠ 未計上")'),
        ("テスト工程",
         f'=IF(SUMIFS($I:$I,$B:$B,"*テスト*",$BH:$BH,"leaf",$A:$A,{NF})'
         f'+SUMIFS($I:$I,$C:$C,"*テスト*",$BH:$BH,"leaf",$A:$A,{NF})>0,{ok},"⚠ 未計上")'),
        ("リリース／ストア申請",
         f'=IF(SUMIFS($I:$I,$B:$B,"*リリース*",$BH:$BH,"leaf",$A:$A,{NF})'
         f'+SUMIFS($I:$I,$D:$D,"*ストア*",$BH:$BH,"leaf",$A:$A,{NF})'
         f'+SUMIFS($I:$I,$D:$D,"*リリース*",$BH:$BH,"leaf",$A:$A,{NF})>0,{ok},"⚠ 未計上")'),
        ("非機能要件チェック",
         f'=IF(COUNTBLANK(\'{NFR_SHEET}\'!$E$4:$E$32)>0,'
         f'"⚠ 未判定 "&COUNTBLANK(\'{NFR_SHEET}\'!$E$4:$E$32)&" 件",'
         f'IF(COUNTIFS(\'{NFR_SHEET}\'!$E$4:$E$32,"対象",\'{NFR_SHEET}\'!$F$4:$F$32,"")'
         f'+COUNTIF(\'{NFR_SHEET}\'!$E$4:$E$32,"要確認")>0,'
         f'"⚠ 未反映/要確認あり（シート参照）","✅ チェック済"))'),
    ]
    for label, formula in checks:
        st(ws.cell(rr, 14), label, F_BASE)
        ws.merge_cells(start_row=rr, start_column=14, end_row=rr, end_column=16)
        st(ws.cell(rr, 17), formula, F_BOLD, align="center")
        rr += 1

    rr += 1
    st(ws.cell(rr, 14),
       "🔍 フェーズ×カテゴリ 工数マトリクス（計上分・人日）— 空欄＝抜け候補 👤", F_BOLD, border=False)
    rr += 1
    cats = []
    for it in flat:
        c = it.get("category", "")
        if it["kind"] == "leaf" and c and c not in cats:
            cats.append(c)
    st(ws.cell(rr, 14), "フェーズ＼カテゴリ", F_HEAD, FILL_HEAD, align="center", wrap=True)
    for j, cat in enumerate(cats):
        st(ws.cell(rr, 15 + j), cat, F_HEAD, FILL_HEAD, align="center", wrap=True)
    ws.row_dimensions[rr].height = 28
    for ph in phases:
        rr += 1
        st(ws.cell(rr, 14), ph, F_BASE)
        for j, cat in enumerate(cats):
            st(ws.cell(rr, 15 + j),
               f'=SUMIFS($G:$G,$B:$B,"{ph}",$C:$C,"{cat}",$BH:$BH,"leaf",$A:$A,{NF})',
               F_BASE, fmt='0.#;;', align="center")
    analysis_last = 15 + max(len(cats), 4)

    # ---- 社内列の折りたたみグループ（J:L と 右側 N.. を隠せる）
    ws.column_dimensions.group("J", "L", outline_level=1, hidden=False)
    ws.column_dimensions.group("N", get_column_letter(analysis_last),
                               outline_level=1, hidden=False)
    ws.sheet_properties.outlinePr.summaryRight = False

    # 顧客提出用の印刷範囲（社内列 J.. を除く A:I）＋横 1 ページに収める
    ws.print_area = f"A1:I{last_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.freeze_panes = ws.cell(header_row + 1, 1)

    st(ws.cell(last_row + 2, 1),
       "⏭ 次にやること: ① 隣の「見積サマリ」で金額を確認 → ② 🔍 抜け漏れ・非機能を ✅ に（右側）→ "
       "③ 👤 利益率を確認 → ④ 🔔 上長・PM のレビュー（Rule 13）→ "
       "⑤ 📤 顧客提出は AI に「提出用の HTML / PDF を生成して」と依頼（or 社内列を隠して A:I を PDF）",
       Font(name=FONT, size=10, bold=True, color=C_PRIMARY), border=False)

    protect(ws)
    return {
        "sheet": name, "row_of": row_of, "flat": flat, "phases": phases,
        "last_row": last_row, "header_row": header_row,
    }


# ---------------------------------------------------------------- 見積サマリ（別シート）
def build_summary(wb, pattern, info, arow):
    """見積サマリ（見積書の隣・数式連携）。案件概要＋金額サマリを 1 枚に。
    金額は見積書の明細を全列 SUMIFS で参照するので、明細を直せば自動で追随する。
    社内金額・利益率は折りたたみグループ＋印刷範囲外（顧客提出は A:D）＝Rule 10。"""
    sname = "見積サマリ" + sheet_suffix(pattern.get("name"))
    ws = wb.create_sheet(sname)
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 3, "B": 30, "C": 13, "D": 16, "E": 3, "F": 15, "G": 12}.items():
        ws.column_dimensions[col].width = w
    d = f"'{info['sheet']}'"
    NF = NOT_OFF

    def esum(col, phase=None):
        p = f',{d}!$B:$B,"{phase}"' if phase else ""
        return (f'SUMIFS({d}!${col}:${col},{d}!$BH:$BH,"leaf",{d}!$A:$A,{NF}{p})'
                f'+SUMIFS({d}!${col}:${col},{d}!$BH:$BH,"pm",{d}!$A:$A,{NF}{p})')
    def dsum(phase=None):
        p = f',{d}!$B:$B,"{phase}"' if phase else ""
        return f'SUMIFS({d}!$G:$G,{d}!$BH:$BH,"leaf",{d}!$A:$A,{NF}{p})'

    pat = f"（{pattern['name']}）" if pattern.get("name") else ""
    st(ws.cell(1, 2), f'="御見積書｜"&設定!$B$4&"{pat}"', F_TITLE, border=False)
    st(ws.cell(2, 2), '=設定!$B$5&" 御中"', F_BOLD, border=False)
    st(ws.cell(3, 2),
       '="作成日: "&TEXT(設定!$B$6,"yyyy/mm/dd")&"　有効期限: "&TEXT(設定!$B$7,"yyyy/mm/dd")'
       '&"　見積精度: "&設定!$B$8&"　Ver."&設定!$B$9', F_MUTED, border=False)
    st(ws.cell(4, 2),
       "このシートは見積書と自動連携。顧客提出は印刷範囲 A:D（F 列＝社内金額・利益率は折りたたみ／出力外）。",
       F_MUTED, border=False)

    # ---- 案件概要
    r = 6
    st(ws.cell(r, 2), "📋 案件概要", F_BOLD, border=False)
    r += 1
    def q(row):  # 前提条件の該当行（空なら控えめ表示）
        return f'=IF(\'前提条件\'!$B{row}="","（前提条件シートに記入）",\'前提条件\'!$B{row})'
    overview = [
        ("案件名", "=設定!$B$4"),
        ("宛先", '=設定!$B$5&" 御中"'),
        ("契約形態", q(arow.get("contract", 0)) if arow.get("contract") else "（前提条件シート参照）"),
        ("対応範囲", q(arow.get("scope_in", 0)) if arow.get("scope_in") else "（前提条件シート参照）"),
        ("主な対象外", q(arow.get("scope_out", 0)) if arow.get("scope_out") else "（前提条件シート参照）"),
        ("作成日 / 有効期限",
         '=TEXT(設定!$B$6,"yyyy/mm/dd")&"　〜　"&TEXT(設定!$B$7,"yyyy/mm/dd")'),
        ("見積精度", "=設定!$B$8"),
    ]
    for label, formula in overview:
        st(ws.cell(r, 2), label, F_BOLD, FILL_SURFACE, wrap=True)
        st(ws.cell(r, 3), formula, F_BASE, wrap=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        r += 1

    # ---- 金額サマリ（見積書と連携）
    r += 1
    st(ws.cell(r, 2), "📊 お見積もりサマリ（見積書と自動連携）", F_BOLD, border=False)
    st(ws.cell(r, 6), "👤 社内金額・利益率は社内限り（Rule 10）", F_MUTED, border=False)
    r += 1
    for c, h in [(2, "区分"), (3, "工数(人日)"), (4, "金額(税抜)"),
                 (6, "社内金額 👤"), (7, "利益率 👤")]:
        st(ws.cell(r, c), h, F_HEAD, FILL_HEAD, align="center")
    r += 1
    row_all = r
    st(ws.cell(r, 2), "プロジェクト全体", F_BOLD, FILL_CHIP)
    st(ws.cell(r, 3), "=" + dsum(), F_BOLD, FILL_CHIP, "0.##", "right")
    st(ws.cell(r, 4), "=" + esum("I"), F_BOLD, FILL_CHIP, YEN)
    st(ws.cell(r, 5), "", F_BOLD, FILL_CHIP)
    st(ws.cell(r, 6), "=" + esum("J"), F_BOLD, FILL_CHIP, YEN)
    st(ws.cell(r, 7), f'=IFERROR((D{r}-F{r})/D{r},"")', F_BOLD, FILL_CHIP, PCT)
    for ph in info["phases"]:
        r += 1
        st(ws.cell(r, 2), ph, F_BASE)
        st(ws.cell(r, 3), "=" + dsum(ph), F_BASE, fmt="0.##", align="right")
        st(ws.cell(r, 4), "=" + esum("I", ph), F_BASE, fmt=YEN)
        st(ws.cell(r, 6), "=" + esum("J", ph), F_BASE, fmt=YEN)
        st(ws.cell(r, 7), f'=IFERROR((D{r}-F{r})/D{r},"")', F_BASE, fmt=PCT)

    # 金額構成（税抜→税込）
    r += 2
    comp_top = r
    comp = [
        ("明細小計（社外）", f"=D{row_all}", F_BASE, None),
        ("リスクバッファ（適用時のみ）",
         f'=IF(バッファ適用="{CHECK_ON}",ROUND(D{row_all}*(バッファ係数-1),0),0)', F_BASE, None),
        ("体制維持費（率×小計）", f"=ROUND(D{row_all}*体制維持費率,0)", F_BASE, None),
        ("お値引き", "=-値引き", F_BASE, None),
        ("ご提示額（税抜）", f"=SUM(D{comp_top}:D{comp_top+3})", F_BOLD, FILL_CHIP),
        ("消費税", None, F_BASE, None),
        ("ご提示額（税込）", f"=D{comp_top+4}+D{comp_top+5}", F_HEAD, FILL_HEAD),
    ]
    for i, (label, formula, font, fill) in enumerate(comp):
        rr = comp_top + i
        big = (label == "ご提示額（税込）")
        if label == "消費税":
            st(ws.cell(rr, 2), '="消費税（"&TEXT(税率,"0%")&"）"', font, fill)
            formula = f"=ROUND(D{comp_top+4}*税率,0)"
        else:
            st(ws.cell(rr, 2), label, font, fill)
        st(ws.cell(rr, 3), "", font, fill)
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)
        st(ws.cell(rr, 4), formula, font, fill, YEN)
    quote_row = comp_top + 4
    tax_row = comp_top + 5
    incl_row = comp_top + 6

    # 採算（社内・利益率）— 印刷範囲(A:D は税込まで)の外に置き、顧客提出物には出さない
    r = incl_row + 2
    st(ws.cell(r, 2), "👤 採算（社内限り・印刷範囲外）", F_BOLD, border=False)
    r += 1
    st(ws.cell(r, 2), "利益率（工数どおり）", F_BASE)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    st(ws.cell(r, 4), f'=IFERROR((D{quote_row}-F{row_all})/D{quote_row},"")', F_BASE, fmt=PCT)
    r += 1
    st(ws.cell(r, 2), "利益率（バッファまで消化）", F_BASE)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    st(ws.cell(r, 4), f'=IFERROR((D{quote_row}-F{row_all}*バッファ係数)/D{quote_row},"")',
       F_BASE, fmt=PCT)

    st(ws.cell(r + 2, 2),
       "⏭ 次にやること: 金額を確認 → 🔔 上長・PM のレビュー（Rule 13）→ "
       "📤 顧客提出は AI に「提出用の HTML / PDF を生成して」と依頼（or A:D を PDF）",
       Font(name=FONT, size=10, bold=True, color=C_PRIMARY), border=False)

    # 社内列（F:G）を折りたたみ／顧客提出の印刷範囲は A:D かつ税込まで（採算は範囲外）
    ws.column_dimensions.group("F", "G", outline_level=1, hidden=False)
    ws.sheet_properties.outlinePr.summaryRight = False
    ws.print_area = f"A1:D{incl_row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    protect(ws)
    q = f"'{sname}'!"
    return {"sheet": sname,
            "days": f"{q}$C${row_all}", "quote_pretax": f"{q}$D${quote_row}",
            "tax": f"{q}$D${tax_row}", "total_incl": f"{q}$D${incl_row}",
            "margin": f"{q}$G${row_all}"}


# ---------------------------------------------------------------- スケジュール
OWNER_COLORS = [("BS", C_PRIMARY), ("共同", C_SKY), ("貴社", C_NAVY),
                ("審査待", C_HAIRLINE), ("リリース", C_SUCCESS)]
GANTT_MONTHS = 18


def phase_days_formula(info, phase):
    """見積明細のフェーズ工数（計上✓のみ）と連動する工数式。
    明細の工数変更・計上－でスケジュールの営業日・総期間も自動で伸縮する。"""
    d = f"'{info['sheet']}'"
    return (f'=ROUND(SUMIFS({d}!$G:$G,{d}!$B:$B,"{phase}",'
            f'{d}!$BH:$BH,"leaf",{d}!$A:$A,{NOT_OFF}),1)')


def default_schedule(info):
    """schedule 未指定時: フェーズごとに 1 行（工数は明細と連動）＋ 受入・申請・リリース行。
    既定は前工程が終わってから次工程（直列＝最長ケース）。並走人数・ラグで短縮する。"""
    rows = [{"task": ph, "owner": "BS", "role": "-", "phase": ph}
            for ph in info["phases"]]
    rows.append({"task": "受入テスト（貴社にて実施）", "owner": "貴社", "role": "-",
                 "days": 0, "biz_days": 10})
    rows.append({"task": "ストア申請・審査待ち", "owner": "審査待", "role": "-",
                 "days": 0, "biz_days": 10})
    rows.append({"task": "リリース（本番公開）", "owner": "リリース", "role": "-",
                 "days": 0, "biz_days": 1})
    return rows


def build_schedule(wb, pattern, info, rates, start_date):
    ws = wb.create_sheet("スケジュール" + sheet_suffix(pattern.get("name")))
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    for col, w in {"B": 12, "C": 11, "D": 10, "E": 9, "F": 8, "G": 9,
                   "H": 11, "I": 11, "J": 9}.items():
        ws.column_dimensions[col].width = w
    for i in range(11, 11 + GANTT_MONTHS):  # 月列
        ws.column_dimensions[get_column_letter(i)].width = 5.5
    last_month_col = get_column_letter(10 + GANTT_MONTHS)

    title = pattern.get("name") or "全体"
    st(ws.cell(1, 1), f"📅 スケジュール（{title}・動的）", F_TITLE, border=False)
    st(ws.cell(2, 1),
       "青いセル（開始日・人数・工数・開始ラグ・審査待ち日数）を変えると、開始/終了日・ガント・総期間が自動で動きます。"
       "工数が数式の工程は見積明細と連動（計上－にすると期間も自動短縮）。"
       "既定は「前の工程が終わったら次を開始」の直列（最長ケース）。並走・前倒しは開始ラグを上書き。",
       F_MUTED, border=False)

    st(ws.cell(4, 1), "⚙ プロジェクト設定", F_BOLD, border=False)
    st(ws.cell(5, 1), "プロジェクト開始日", F_BASE)
    st(ws.cell(5, 2), start_date, F_BASE, FILL_INPUT, "yyyy/mm/dd", unlock=True)
    st(ws.cell(6, 1), "営業日/月", F_BASE)
    st(ws.cell(6, 2), "=営業日月", F_BASE, fmt="0")

    st(ws.cell(8, 1), "👥 並走人数（役割ごと）", F_BOLD, border=False)
    st(ws.cell(9, 1), "役割", F_HEAD, FILL_HEAD, align="center")
    st(ws.cell(9, 2), "人数", F_HEAD, FILL_HEAD, align="center")
    role_top = 10
    role_keys = [x["key"] for x in rates] + ["-"]
    for i, key in enumerate(role_keys):
        st(ws.cell(role_top + i, 1), key, F_BASE)
        st(ws.cell(role_top + i, 2), 1, F_BASE, FILL_INPUT, "0", unlock=True)
    role_end = role_top + len(role_keys) - 1
    st(ws.cell(role_end + 1, 1), "※「-」＝役割未指定の工程（人数はこの行を使用）",
       F_MUTED, border=False)

    table_top = role_end + 3
    st(ws.cell(table_top - 1, 1), "📅 スケジュール表（自動計算）", F_BOLD, border=False)
    month_row = table_top      # EDATE 行
    head_row = table_top + 1
    for m in range(GANTT_MONTHS):
        c = ws.cell(month_row, 11 + m)
        st(c, f'=EDATE(DATE(YEAR($B$5),MONTH($B$5),1),{m})', F_MUTED,
           fmt='yy/m', align="center")
    headers = ["工程", "担当", "役割", "工数(人日)", "並走人数", "営業日",
               "開始ラグ(営業日)", "開始日", "終了日", "必要月数"]
    for i, h in enumerate(headers, start=1):
        st(ws.cell(head_row, i), h, F_HEAD, FILL_HEAD, align="center", wrap=True)
    for m in range(GANTT_MONTHS):
        st(ws.cell(head_row, 11 + m),
           f'=TEXT({get_column_letter(11 + m)}{month_row},"m月")',
           F_HEAD, FILL_HEAD, align="center")
    ws.row_dimensions[head_row].height = 28

    dv_owner = DataValidation(type="list", formula1='"BS,共同,貴社,審査待,リリース"',
                              allow_blank=True)
    dv_role = DataValidation(type="list",
                             formula1=f"=$A${role_top}:$A${role_end}", allow_blank=True)
    ws.add_data_validation(dv_owner)
    ws.add_data_validation(dv_role)

    sched = pattern.get("schedule") or default_schedule(info)
    r = head_row
    first_task_row = head_row + 1
    for i, task in enumerate(sched):
        r += 1
        st(ws.cell(r, 1), task["task"], F_BASE, wrap=True, unlock=True)
        st(ws.cell(r, 2), task.get("owner", "BS"), F_BASE, align="center", unlock=True)
        dv_owner.add(ws.cell(r, 2))
        st(ws.cell(r, 3), task.get("role", "-"), F_BASE, FILL_INPUT,
           align="center", unlock=True)
        dv_role.add(ws.cell(r, 3))
        # "phase" 指定の工程は見積明細のフェーズ工数と連動（動的）。それ以外は青セル入力
        if task.get("phase"):
            days_val = phase_days_formula(info, task["phase"])
        else:
            days_val = task.get("days", 0)
        st(ws.cell(r, 4), days_val, F_BASE,
           None if isinstance(days_val, str) else FILL_INPUT,
           "0.#", "right", unlock=True)
        st(ws.cell(r, 5),
           f'=MAX(1,IFERROR(VLOOKUP($C{r},$A${role_top}:$B${role_end},2,FALSE),1))',
           F_BASE, fmt="0", align="right")
        if task.get("biz_days"):  # 審査待ちなど工数ゼロで期間だけある行
            st(ws.cell(r, 6), task["biz_days"], F_BASE, FILL_INPUT, "0", "right",
               unlock=True)
        else:
            st(ws.cell(r, 6),
               f'=IF(OR($D{r}=0,$D{r}="",$D{r}="－"),0,ROUNDUP($D{r}/$E{r},0))',
               F_BASE, fmt="0", align="right")
        lag = task.get("lag")
        if lag is None:
            lag = 0 if i == 0 else f"=$G{r - 1}+$F{r - 1}"
        st(ws.cell(r, 7), lag, F_BASE, FILL_INPUT, "0", "right", unlock=True)
        st(ws.cell(r, 8), f'=WORKDAY($B$5-1,$G{r}+1)', F_BASE, fmt='m/d')
        st(ws.cell(r, 9), f'=IF($F{r}<=1,$H{r},WORKDAY($H{r},$F{r}-1))',
           F_BASE, fmt='m/d')
        st(ws.cell(r, 10), f'=IF($F{r}=0,"－",ROUND($F{r}/$B$6,1)&"ヶ月")',
           F_BASE, align="right")
        for m in range(GANTT_MONTHS):
            st(ws.cell(r, 11 + m), "", F_BASE)
    last_task_row = r

    # ガント（担当別色・条件付き書式）
    rng = f"K{first_task_row}:{last_month_col}{last_task_row}"
    for owner, color in OWNER_COLORS:
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'AND($B{first_task_row}="{owner}",'
                     f'$H{first_task_row}<=EOMONTH(K${month_row},0),'
                     f'$I{first_task_row}>=K${month_row})'],
            fill=PatternFill("solid", start_color=color), stopIfTrue=False))

    r += 2
    st(ws.cell(r, 1), "🏁 プロジェクト終了日", F_BOLD)
    st(ws.cell(r, 8), f"=MAX(I{first_task_row}:I{last_task_row})", F_BOLD, fmt='yyyy/m/d')
    r += 1
    st(ws.cell(r, 1), "総期間（営業日）", F_BASE)
    st(ws.cell(r, 8), f"=NETWORKDAYS($B$5,H{r - 1})", F_BASE, fmt="0")
    r += 1
    st(ws.cell(r, 1), "総期間（カ月）", F_BASE)
    st(ws.cell(r, 8), f'=ROUND(H{r - 1}/$B$6,1)&"ヶ月"', F_BASE, align="right")
    r += 2
    st(ws.cell(r, 1), "凡例: ", F_MUTED, border=False)
    for i, (owner, color) in enumerate(OWNER_COLORS):
        c = ws.cell(r, 2 + i)
        c.fill = PatternFill("solid", start_color=color)
        st(ws.cell(r + 1, 2 + i), owner, F_MUTED, border=False, align="center")
    st(ws.cell(r, 8),
       f"※ ガントは {GANTT_MONTHS} ヶ月分。超える場合は {last_month_col} 列を右へコピー",
       F_MUTED, border=False)
    st(ws.cell(r + 3, 1),
       "⏭ 次にやること: 総期間とリリース日を確認 → お客様向けの月次ガントは AI に"
       "「提出用スケジュール（HTML）を生成して」と依頼",
       Font(name=FONT, size=10, bold=True, color=C_PRIMARY), border=False)
    protect(ws)
    return ws


# ---------------------------------------------------------------- 案比較（複数案のみ）
def build_comparison(wb, infos):
    """複数案（Phase 別・スコープ別）の比較表。各案の明細シートと数式連動。"""
    ws = wb.create_sheet("案比較")
    ws.sheet_view.showGridLines = False
    widths = {"A": 22, "B": 12, "C": 16, "D": 14, "E": 16, "F": 11, "G": 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    st(ws.cell(1, 1), "🗂 案比較（自動計算・各案の明細と連動）", F_TITLE, border=False)
    st(ws.cell(2, 1), "スコープ・Phase の違う複数案を並べて比較。値引き相談では案の切替（スコープ調整）を第一候補に（Rule 5）。",
       F_MUTED, border=False)
    headers = ["案", "工数(計上分)", "ご提示額(税抜)", "消費税", "ご提示額(税込)",
               "利益率 👤", "メモ（この案の範囲）"]
    for i, h in enumerate(headers, start=1):
        st(ws.cell(3, i), h, F_HEAD, FILL_HEAD, align="center")
    r = 3
    for info in infos:
        r += 1
        c = info["cells"]  # 見積サマリの各セル（シート名込みの絶対参照）
        st(ws.cell(r, 1), info["sheet"].replace("見積書_", "").replace("見積書", "本見積もり"),
           F_BOLD)
        st(ws.cell(r, 2), f"={c['days']}", F_BASE, fmt="0.##", align="right")
        st(ws.cell(r, 3), f"={c['quote_pretax']}", F_BASE, fmt=YEN)
        st(ws.cell(r, 4), f"={c['tax']}", F_BASE, fmt=YEN)
        st(ws.cell(r, 5), f"={c['total_incl']}", F_BOLD, FILL_CHIP, YEN)
        st(ws.cell(r, 6), f"={c['margin']}", F_BASE, fmt=PCT)
        st(ws.cell(r, 7), "", F_BASE, FILL_INPUT, wrap=True, unlock=True)
    st(ws.cell(r + 2, 1),
       "⏭ 次にやること: お客様に提示する案を決めて 🔔 レビューへ（利益率 👤 は社外に出さない）",
       Font(name=FONT, size=10, bold=True, color=C_PRIMARY), border=False)
    protect(ws)
    return ws


# ---------------------------------------------------------------- ヒアリングシート（観点チェック）
# (カテゴリ, 項目, 内容・目的)。人は「確認」列に ✓ を付けるだけ／AI が分かる項目は回答を事前記入。
HEARING_ITEMS = [
    ("プロジェクト基本情報", "クライアント名", "【内容】正式名称、担当部署\n【目的】見積書や契約書の宛先として必要。"),
    ("プロジェクト基本情報", "アプリ名称(仮)", "【内容】プロジェクトの仮称または正式名称\n【目的】共通言語として使用するため。"),
    ("プロジェクト基本情報", "システム化の背景", "【内容】なぜこのプロジェクトが必要になったのか。現状の課題。\n【目的】プロジェクトの根本的な目的を理解し、最適な解決策を提案する。"),
    ("プロジェクト基本情報", "プロジェクトの目的とゴール", "【内容】このプロジェクトで達成したい具体的な成果(売上UP、効率化など)。\n【目的】機能の優先順位付けや、成功を定義する上での基準とする。"),
    ("プロジェクト基本情報", "RFP(提案依頼書)の有無", "【内容】プロジェクトの目的や要件がまとまった公式な資料の有無。\n【目的】要求事項の明確度を判断し、初期の要件整理工数を見積もるため。"),
    ("プロジェクト基本情報", "技術的な実現可能性の調査(PoC)の要否", "【内容】AI、画像解析、特殊なハードウェア連携など、技術的に不確実性の高い機能はあるか。\n【目的】本開発とは別で、小規模な試作(PoC)による実現性検証が必要か判断する。本開発の見積もりとは別の工数・費用となるため、最初に確認が必須。"),
    ("成功指標と計測", "成功指標(KGI/KPI)", "【内容】このプロジェクトが成功したと判断する具体的な数値目標と達成時期\n【目的】合意した達成基準を元に、機能の優先順位やスコープを判断・調整するため。「作って終わり」を防ぐ。"),
    ("成功指標と計測", "アクセス解析ツールの導入", "【内容】利用状況の分析(どの画面がよく見られているか、など)のために、専用のツールを導入するか。\n【目的】解析ツールのSDK組み込みや、計測したいイベント(ボタンクリックなど)の設計・実装に追加の工数がかかるため。"),
    ("体制とガバナンス", "クライアント側専任担当者", "【内容】プロジェクトの窓口となる担当者の有無。\n【目的】専任担当者がいない場合、コミュニケーションコストや意思決定の遅延が予想され、工数に影響する。"),
    ("体制とガバナンス", "意思決定プロセス", "【内容】仕様変更や追加費用の承認は、誰が、どのようなプロセス(稟議など)で行うか。\n【目的】承認階層やプロセスが複雑な場合、手戻りや意思決定の遅延がスケジュールに大きく影響するため、事前に把握しておく必要がある。"),
    ("スコープ(機能要件)", "機能一覧(マスト/ウォント)", "【内容】実装したい機能の洗い出し。必須機能と希望機能を分けておく。\n【目的】スコープ確定・見積もりの基礎とする。"),
    ("スコープ(機能要件)", "管理画面の有無", "【内容】運営者が使用する管理機能の要否。\n【目的】「有」の場合、別アプリ相当の設計・開発・テスト工数が追加で必要になるため。"),
    ("スコープ(機能要件)", "アプリ内課金(IAP)/決済", "【内容】App Store/Google Play経由での課金や、クレジットカード決済などが必要か。\n【目的】決済機能は審査が厳しく、返金ポリシーなどの運用ルール設計も必要になるため、プロジェクトの難易度と工数が大幅に増加する。"),
    ("スコープ(機能要件)", "マネタイズの軸", "【内容】アプリ内課金、広告、サブスクリプション、EC連携など、このサービスでどのように収益を上げる計画か。\n【目的】収益化の方法は、決済機能の実装だけでなく、ユーザー導線やUI設計の根幹に関わるため。"),
    ("スコープ(機能要件)", "CSVエクスポート機能の有無", "【内容】管理画面などから、データをCSVファイルとしてダウンロードする機能は必要か。\n【目的】管理画面の「あるある機能」であり、データ抽出ロジックとファイル生成処理の実装が必要になるため、事前に確認が必要。"),
    ("スコープ(機能要件)", "データ移行の有無", "【内容】既存システムから新システムへデータ(例：会員情報)を移行する必要があるか。\n【目的】データの抽出、変換、投入(ETL)の設計とリハーサルに大きな工数がかかるため。"),
    ("スコープ(機能要件)", "端末権限の利用", "【内容】位置情報、カメラ、プッシュ通知など、スマートフォンの機能を利用するか。\n【目的】各機能の利用許可を求めるUI設計と、ユーザーが拒否した場合の代替フローを設計する必要があるため。"),
    ("スコープ(非機能要件)", "対応OS・デバイス", "【内容】iOS/Android/PC/タブレットなど、どのデバイスで利用できるようにするか。\n【目的】対応デバイスが増えるほど、それぞれの画面サイズに合わせた設計・実装・テスト工数が増加する。"),
    ("スコープ(非機能要件)", "サポート範囲(OS/ブラウザ)", "【内容】各OSやブラウザのどのバージョンまでサポートするか。\n【目的】サポート範囲が広い(古いバージョンを含む)ほど、テスト工数や古い環境への対応コストが増加する。"),
    ("スコープ(非機能要件)", "性能目標", "【内容】「〇〇秒以内に表示」「同時に〇〇人がアクセスしても耐えられる」といった、速度や負荷に関する具体的な目標はあるか。\n【目的】高い性能目標を達成するには、それに合わせたインフラ設計やアーキテクチャ選定が必要になり、コストに大きく影響するため。"),
    ("スコープ(非機能要件)", "オフライン対応の有無", "【内容】電波がない状態でも、一部機能が利用できるようにするか。\n【目的】端末内にデータを保存・同期する複雑なロジックが必要になり、工数が大幅に増加するため。"),
    ("スコープ(非機能要件)", "画面の向き(縦固定/横対応)", "【内容】スマートフォンやタブレットを横向きにした際の表示に対応するか。\n【目的】横向きは、レイアウトが崩れないように設計・実装する必要があるため、工数に直接影響する。「縦向き固定」と割り切るだけで、工数を削減できる。"),
    ("スコープ(非機能要件)", "多言語対応の有無", "【内容】日本語以外に対応するか。対応する場合、どの言語か。\n【目的】翻訳テキストの管理、言語ごとのレイアウト調整(文字長の変動)など、工数への影響が非常に大きいため。"),
    ("スコープ(非機能要件)", "計画停止の有無", "【内容】24時間365日稼働が前提か、深夜帯などに計画的なメンテナンス停止が可能か。\n【目的】無停止でのリリース(デプロイ)が求められる場合、サーバー構成やリリース手順が複雑になり、工数が増加するため。"),
    ("デザイン・体制", "デザイン作成の担当", "【内容】デザイン(UI/UX)はどちらが用意するか(顧客支給 or 自社制作)。\n【目的】自社でUIデザイン、UX設計から行う場合、工数が大幅に増加する。"),
    ("デザイン・体制", "参考にしているサービス", "【内容】デザインや機能のイメージに近い、参考となるアプリやWebサイト。\n【目的】お客様の頭の中にある完成イメージを具体的に共有してもらうことで、デザインの方向性を固め、手戻りを防ぐ。"),
    ("デザイン・体制", "デザインのテーマカラー", "【内容】ブランドカラーや、希望するメインカラー、アクセントカラー。\n【目的】デザインの基調となる色を早期に決定することで、デザイン制作をスムーズに進めるため。"),
    ("デザイン・体制", "フォントの指定", "【内容】OS標準のフォント以外に、特定のブランドフォントや有償フォントを使用するか。\n【目的】有償フォントを利用する場合、ライセンス費用の確認と、アプリへの組み込み工数が別途発生する。"),
    ("デザイン・体制", "スプラッシュ画面のアニメーション有無", "【内容】アプリ起動時に表示される画面(スプラッシュ)に、単純なロゴ表示だけでなく、アニメーションを実装するか。\n【目的】アニメーションは、デザイン工数と実装工数の両方が追加で発生する。ブランディングに関わる重要な要素だが、コスト要因でもあるため確認が必要。"),
    ("デザイン・体制", "素材の提供の有無", "【内容】アプリ内で使用する画像、テキスト、動画などの素材は誰が用意するか。\n【目的】素材(特に原稿)の提供が遅れると、プロジェクト全体の遅延に直結する。"),
    ("デザイン・体制", "アクセシビリティ対応", "【内容】公共機関の案件などで、WCAGなどのアクセシビリティ基準に準拠する必要があるか。\n【目的】音声読み上げ対応や、十分なコントラスト比の確保など、専門的な設計・実装・テストが必要になり、工数が大幅に増加するため。"),
    ("インフラ・セキュリティ", "個人情報の取り扱いの有無", "【内容】氏名、住所、電話番号、決済情報などを扱うか。\n【目的】「有」の場合、セキュリティ要件が格段に上がり、暗号化やアクセス制御など厳重な設計・実装が必要。"),
    ("インフラ・セキュリティ", "権限管理・認可設計", "【内容】ユーザーの種類(管理者、一般ユーザー等)によって、使える機能や見られるデータを制御する必要があるか。\n【目的】権限管理は設計・実装が複雑になりがちで、工数に大きく影響するため。"),
    ("インフラ・セキュリティ", "監査ログ・操作ログの要否", "【内容】「誰が、いつ、何をしたか」といった操作履歴を保存する必要があるか。\n【目的】ログの保存要件(何を、どのくらいの期間)は、サーバーのストレージコストや、ログ収集・解析基盤の構築工数に直接影響するため。"),
    ("インフラ・セキュリティ", "脆弱性診断の実施有無", "【内容】第三者機関によるセキュリティの専門的な検査(脆弱性診断)を実施するか。\n【目的】診断の実施費用(実費)と、検出された脆弱性の改修対応に大きな工数がかかる可能性があるため。"),
    ("インフラ・セキュリティ", "バックアップ/DR計画", "【内容】災害などでシステムが停止した場合、どのくらいの時間で復旧させる必要があるか。(RTO/RPO)\n【目的】高い事業継続性を求める場合、バックアップやDR(災害復旧)サイトの構築・維持に専門的な工数とインフラコストがかかるため。"),
    ("インフラ・セキュリティ", "ドメインの有無", "【内容】APIサーバーやWebサイトで利用するドメインは取得済みか。新規取得・設定が必要か。\n【目的】ドメインの取得・設定作業の有無を確認するため。ドメインはお客様の資産であり、事前に取得方法や所有者を決めておく必要がある。"),
    ("リリース・運用", "配布形態", "【内容】ストア公開か、クローズド配布(MDM, In-House)か。\n【目的】一般公開(ストア申請)は審査プロセスが必須。クローズド配布は別の仕組み構築が必要。"),
    ("リリース・運用", "テストの配信方法", "【内容】開発中のアプリを関係者にどう配布・確認するか(TestFlight, Firebase App Distributionなど)。\n【目的】配布方法のセットアップや管理に工数がかかる。また、顧客側の確認環境(テスト端末の有無など)を把握するため。"),
    ("リリース・運用", "強制アップデートの有無", "【内容】古いバージョンのアプリを強制的に使えなくする仕組みが必要か。\n【目的】全ユーザーの利用バージョンを統一するための重要な機能。バージョン管理APIや画面の実装が追加で必要になるため。"),
    ("リリース・運用", "メンテナンスモードの有無", "【内容】サーバーメンテナンス中などに、アプリ側で「メンテナンス中」と表示する機能は必要か。\n【目的】「有」の場合、メンテナンス状態を管理する仕組みや専用画面の実装が必要になるため、追加の工数が発生する。"),
    ("リリース・運用", "ストア掲載画像の作成有無", "【内容】App StoreやGoogle Playに掲載するスクリーンショットや紹介画像をどちらが作成するか。\n【目的】アプリのダウンロード率に直結する重要な要素であり、誰が、いつまでに用意するのかを明確にしておかないと、リリース直前に慌てることになる。"),
    ("リリース・運用", "保守・運用体制", "【内容】リリース後の不具合対応や、サーバー監視、OSアップデート対応をどうするか。\n【目的】リリース後の保守運用契約が別途必要なのか、自社で対応するのかを明確にする。"),
    ("成果物・制約", "納品物(成果物)の定義", "【内容】設計書・仕様書の提出は必要か。どのレベルの詳細さまで必要か。\n【目的】ドキュメント作成は工数に直結する。必要なドキュメントの種類と粒度を事前に合意する。"),
    ("成果物・制約", "ソースコード引渡義務の有無", "【内容】開発したソースコード全ての引き渡しが必要か。\n【目的】著作権やライセンスの取り決め、コードの清書(可読性担保)など、契約と工数に影響する。"),
    ("成果物・制約", "契約上の特記事項", "【内容】ソースコードの知的財産権の扱いや、再委託の可否など、契約面で特別な条件はあるか。\n【目的】知財の完全譲渡や再委託の禁止などは、開発体制や利用技術の選定に制約を与え、見積もりに影響するため。"),
    ("成果物・制約", "法務/規制の確認", "【内容】資金決済法、薬機法、景表法など、事業分野に特有の法律やガイドラインはあるか。\n【目的】特定の業法に対応する場合、専門的な知識が必要となり、審査や表示ルールの対応で工数が増加するため。"),
    ("成果物・制約", "利用規約・プライバシーポリシーの作成", "【内容】アプリ内で表示・同意取得が必要な利用規約やプライバシーポリシーはどちらが用意するか。\n【目的】ストア申請や個人情報取り扱いに必須のドキュメント。作成には法務確認が必要な場合が多く、責任の所在を明確にする必要がある。"),
    ("成果物・制約", "完成後の拡張(将来的な展望)", "【内容】今回のリリース後、1年後、3年後にどのような機能追加や事業展開を考えているか。\n【目的】将来的な拡張性(スケーラビリティ)を考慮してシステムを設計するかどうかで、初期のアーキテクチャ選定や工数が変わるため。"),
    ("成果物・制約", "スケジュールの緊急性", "【内容】「いつまでに」という絶対的な期限(例：イベント開催日)があるか。\n【目的】期限が厳しい場合は、投入リソースを増やしたり、スコープを削るなどの調整が必要になる。"),
    ("成果物・制約", "予算規模", "【内容】大まかな予算感はどれくらいか。\n【目的】予算に応じて、実現可能な機能範囲や技術選定、体制を逆算して提案する必要がある。"),
    ("体制・外部連携（追加確認）", "オフショア", "【内容】オフショア開発（ベトナム/北京 等）を利用するか。\n【目的】単価・コミュニケーション体制・品質管理・ブリッジSE の要否に影響。見積もりの単価前提そのものに関わる。"),
    ("体制・外部連携（追加確認）", "外部開発者との協業", "【内容】外部パートナー・フリーランス・他社と協業して開発するか。\n【目的】体制・契約（再委託）・情報共有・レビュー体制の設計に影響。"),
    ("体制・外部連携（追加確認）", "SDK等の提供・利用", "【内容】お客様や第三者から提供される SDK・ライブラリ・API を組み込むか（または当方が提供するか）。\n【目的】ライセンス確認・組み込み工数・依存/バージョン管理・サポート範囲に影響。"),
]


def derive_hearing_answers(data):
    """見積もりデータから自明なヒアリング回答を推定して事前記入する（人の確認負荷を下げる）。
    例: オフショア単価が使われていれば「オフショア利用＝有」を自動で埋める。"""
    rates = data.get("rates", [])
    rate_join = " ".join((str(r.get("key", "")) + str(r.get("name", ""))) for r in rates)
    roles, texts = set(), []
    for pat in data.get("patterns", []):
        def walk(items):
            for it in items:
                if it.get("type") == "parent":
                    texts.append(str(it.get("item", "")) + str(it.get("desc", "")))
                    walk(it.get("children", []))
                else:
                    roles.add(str(it.get("role", "")))
                    texts.append(str(it.get("item", "")) + str(it.get("desc", ""))
                                 + str(it.get("category", "")) + str(it.get("impl", "")))
        walk(pat.get("items", []))
    T = " ".join(texts)
    a = data.get("assumptions") or {}
    def joinval(v):
        return " ".join(map(str, v)) if isinstance(v, list) else str(v or "")
    scope_out = joinval(a.get("scope_out"))
    env = joinval(a.get("environment"))

    ans = {}
    offshore = ("オフショア" in roles) or any(
        k in rate_join for k in ["ベトナム", "北京", "オフショア", "offshore", "Offshore", "OFFSHORE"])
    ans["オフショア"] = ("利用する（オフショア単価を見積に計上済み）" if offshore
                     else "利用しない想定（国内単価で計上）")
    if any(k in T for k in ["管理画面", "CMS"]):
        ans["管理画面の有無"] = "有（管理画面を見積に計上）"
    if any(k in T for k in ["決済", "課金", "IAP", "サブスク", "サブスクリプション"]):
        ans["アプリ内課金(IAP)/決済"] = "有（決済/課金関連を計上）"
    if "多言語" in T:
        ans["多言語対応の有無"] = "有（多言語対応を計上）"
    elif "多言語" in scope_out:
        ans["多言語対応の有無"] = "対象外（別途お見積もり）"
    if any(k in T for k in ["プッシュ", "通知", "位置情報", "カメラ", "権限"]):
        ans["端末権限の利用"] = "利用あり（例: プッシュ通知等を計上）"
    if any(k in T for k in ["認証", "ログイン", "個人情報", "セキュリティ", "本人確認"]):
        ans["個人情報の取り扱いの有無"] = "取り扱いあり想定（認証・セキュリティ設計を計上）"
    if "脆弱性" in T:
        ans["脆弱性診断の実施有無"] = "実施（脆弱性診断を計上）"
    if any(k in T for k in ["データ移行", "移行", "ETL"]):
        ans["データ移行の有無"] = "有（データ移行を計上）"
    design = any(("デザイン" in str(it.get("phase", "")) or "デザイン" in str(it.get("category", "")))
                 for pat in data.get("patterns", []) for it in pat.get("items", []))
    if design:
        ans["デザイン作成の担当"] = "自社制作想定（UI/UXデザインを見積に計上）"
    if env:
        ans["対応OS・デバイス"] = "前提条件シート参照（" + env[:36] + ("…" if len(env) > 36 else "") + "）"
    return ans


def build_hearing(wb, data):
    """ヒアリングシート（観点チェック）。人は「確認」に ✓ を付けるだけ／AI が既知の回答を事前記入。
    ✓ を付けると行の色が変わる（観点漏れ防止）。全項目に回答する必要はない。"""
    ws = wb.create_sheet("ヒアリングシート")
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 7, "B": 5, "C": 17, "D": 26, "E": 74, "F": 30}.items():
        ws.column_dimensions[col].width = w

    ans = derive_hearing_answers(data)
    n = len(HEARING_ITEMS)
    first, last = 5, 4 + n

    st(ws.cell(1, 1), "🗣 ヒアリングシート（観点チェック）— 人は「確認」に ✓ を付けるだけ",
       F_TITLE, border=False)
    st(ws.cell(2, 1),
       "見積もりから分かる項目は AI が回答を事前記入済み（水色セル＝要確認・修正可）。"
       "全項目に回答する必要はありません。観点漏れ防止のため『確認』列に ✓ だけ付けてください（✓ で行の色が変わります）。",
       F_MUTED, border=False)
    st(ws.cell(3, 6),
       f'="確認済み "&COUNTIF($A${first}:$A${last},"✓")&" / {n} 項目"',
       Font(name=FONT, size=11, bold=True, color=C_PRIMARY), align="right")

    headers = ["確認", "#", "カテゴリ", "項目", "内容・目的", "回答（分かる範囲でOK）"]
    for i, h in enumerate(headers, start=1):
        st(ws.cell(4, i), h, F_HEAD, FILL_HEAD, align="center")
    ws.freeze_panes = "A5"

    dv_chk = DataValidation(type="list", formula1='"✓"', allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_chk)

    prev_cat = None
    group_start = None
    cat_merges = []
    r = 4
    for idx, (cat, item, naiyou) in enumerate(HEARING_ITEMS, start=1):
        r += 1
        st(ws.cell(r, 1), "", F_BASE, align="center", unlock=True)
        dv_chk.add(ws.cell(r, 1))
        st(ws.cell(r, 2), idx, F_MUTED, align="center")
        # カテゴリ: 同じ値が続く間はグループの先頭セルにだけ書き、あとで縦結合する
        if cat != prev_cat:
            if group_start is not None and r - 1 > group_start:
                cat_merges.append((group_start, r - 1))
            group_start = r
            st(ws.cell(r, 3), cat, F_BOLD, wrap=True)
        else:
            st(ws.cell(r, 3), "", F_BOLD, wrap=True)
        prev_cat = cat
        st(ws.cell(r, 4), item, F_BASE, wrap=True)
        st(ws.cell(r, 5), naiyou, F_MUTED, wrap=True)
        pre = ans.get(item, "")
        st(ws.cell(r, 6), pre, F_BASE, (FILL_INPUT if pre else None),
           wrap=True, unlock=True)
        ws.row_dimensions[r].height = 46
    if group_start is not None and r > group_start:
        cat_merges.append((group_start, r))
    # カテゴリ列を縦結合（同カテゴリの行をまとめる）。結合セルは上寄せ中央で見やすく
    for a, b in cat_merges:
        ws.merge_cells(start_row=a, start_column=3, end_row=b, end_column=3)
        ws.cell(a, 3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ✓ を付けた行を薄い緑にハイライト（色が変わる）
    ws.conditional_formatting.add(
        f"A{first}:F{last}",
        FormulaRule(formula=[f'$A{first}="✓"'],
                    fill=PatternFill("solid", start_color="E4F6EC")))
    # 確認セル自体は ✓ を緑太字で強調
    ws.conditional_formatting.add(
        f"A{first}:A{last}",
        FormulaRule(formula=[f'$A{first}="✓"'],
                    font=Font(name=FONT, size=12, bold=True, color=C_SUCCESS)))

    st(ws.cell(last + 2, 1),
       "⏭ 次にやること: ① 上から観点を確認し『確認』に ✓（未確認の観点をゼロに）→ "
       "② 追加で分かった条件は回答欄・前提条件シート・見積書に反映 → ③ 未確定は Rule 8 で TBD/別途明記",
       Font(name=FONT, size=10, bold=True, color=C_PRIMARY), border=False)
    protect(ws)
    return ws


# ---------------------------------------------------------------- 非機能要件チェック
def build_nfr(wb, overrides):
    """非機能要件チェックシート（Rule 2 / Step 03）。
    判定=対象 にしたら F 列に見積明細のどの項目で工数化したかを書く。
    未反映・要確認・未判定は自動で警告（見積明細の抜け漏れチェックにも連動）。"""
    ws = wb.create_sheet(NFR_SHEET)
    ws.sheet_view.showGridLines = False
    widths = {"A": 5, "B": 18, "C": 52, "D": 14, "E": 10, "F": 34, "G": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    st(ws.cell(1, 1), "🛡 非機能要件チェック（Rule 2: 機能だけ数えて非機能を抜かさない）",
       F_TITLE, border=False)
    st(ws.cell(2, 1),
       '="判定の進み具合: 対象 "&COUNTIF($E$4:$E$32,"対象")&" 件（うち明細未反映 "'
       '&COUNTIFS($E$4:$E$32,"対象",$F$4:$F$32,"")&" 件）／ 対象外 "'
       '&COUNTIF($E$4:$E$32,"対象外")&" 件 ／ 要確認 "&COUNTIF($E$4:$E$32,"要確認")'
       '&" 件 ／ 未判定 "&COUNTBLANK($E$4:$E$32)&" 件　※ 最低 C(セキュリティ)・D(運用)・'
       'G1(ストア審査) は必ず判断（抜けやすい筆頭）"',
       F_MUTED, border=False)

    headers = ["#", "区分", "チェック観点", "対象分野", "判定",
               "見積明細への反映（項目名）", "メモ（目標値・理由）"]
    for i, h in enumerate(headers, start=1):
        st(ws.cell(3, i), h, F_HEAD, FILL_HEAD, align="center")
    ws.cell(3, 5).comment = Comment(
        "対象＝見積明細に工数を積む（F列に項目名を書く）／対象外＝提出物の前提に明記（Rule 3）／"
        "要確認＝❓でお客様・社内に確認（Rule 8）", "estimation-kit")

    dv_field = DataValidation(
        type="list",
        formula1='"アプリ,Web・フロント,API・サーバ,インフラ,管理画面,外部連携,共通,アプリ/Web,サーバ・インフラ"',
        allow_blank=True, showErrorMessage=False)
    dv_status = DataValidation(type="list", formula1='"対象,対象外,要確認"',
                               allow_blank=True)
    ws.add_data_validation(dv_field)
    ws.add_data_validation(dv_status)

    ov = {o["code"]: o for o in (overrides or [])}
    r = 3
    for code, group, item, field in NFR_DEFAULTS:
        r += 1
        o = ov.get(code, {})
        st(ws.cell(r, 1), code, F_BOLD, align="center")
        st(ws.cell(r, 2), group, F_MUTED, wrap=True)
        st(ws.cell(r, 3), item, F_BASE, wrap=True)
        st(ws.cell(r, 4), o.get("field", field), F_BASE, FILL_INPUT,
           align="center", unlock=True)
        st(ws.cell(r, 5), o.get("status", ""), F_BASE, FILL_INPUT,
           align="center", unlock=True)
        st(ws.cell(r, 6), o.get("reflect", ""), F_BASE, FILL_INPUT,
           wrap=True, unlock=True)
        st(ws.cell(r, 7), o.get("note", ""), F_MUTED, FILL_INPUT,
           wrap=True, unlock=True)
        dv_field.add(ws.cell(r, 4))
        dv_status.add(ws.cell(r, 5))
    last = r

    # 対象なのに明細未反映 → 行を赤く / 要確認 → 判定セルを強調
    ws.conditional_formatting.add(
        f"A4:G{last}",
        FormulaRule(formula=[f'AND($E4="対象",$F4="")'],
                    fill=PatternFill("solid", start_color=C_ERROR_BG)))
    ws.conditional_formatting.add(
        f"E4:E{last}",
        FormulaRule(formula=[f'$E4="要確認"'],
                    font=Font(name=FONT, size=10, bold=True, color=C_ERROR)))

    st(ws.cell(last + 2, 1),
       "⏭ 次にやること: 「対象」にした観点は見積明細に工数を積んで F 列に項目名を記入 → "
       "「対象外」は提出物の前提（対象外）に明記 → 「要確認」は AI に ❓ で確認を依頼",
       Font(name=FONT, size=10, bold=True, color=C_PRIMARY), border=False)
    protect(ws)
    return ws


# ---------------------------------------------------------------- 機能×API 対応表
def build_feature_api(wb, rows):
    ws = wb.create_sheet("機能×API対応表")
    headers = ["機能領域", "詳細項目", "利用API（タグ）", "API数",
               "Phase1", "Phase2", "フル", "実装方式", "備考"]
    widths = [16, 34, 30, 8, 8, 8, 8, 26, 30]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        st(ws.cell(1, i), h, F_HEAD, FILL_HEAD, align="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    sample = rows or [
        {"area": "（例）認証", "detail": "ログイン（JWT）", "apis": "認証",
         "api_count": 3, "p1": 14, "p2": 0, "full": 14,
         "impl": "ネイティブ", "note": ""},
        {"area": "（例）買取・譲渡", "detail": "査定・買取・譲渡", "apis": "買取/譲渡",
         "api_count": 30, "p1": 0, "p2": 0, "full": 0,
         "impl": "★スコープ未確定", "note": "要件確定後に別途（0計上で一覧に残す）"},
    ]
    r = 1
    for row in sample:
        r += 1
        vals = [row.get("area", ""), row.get("detail", ""), row.get("apis", ""),
                row.get("api_count", ""), row.get("p1", ""), row.get("p2", ""),
                row.get("full", ""), row.get("impl", ""), row.get("note", "")]
        for i, v in enumerate(vals, start=1):
            st(ws.cell(r, i), v, F_BASE, FILL_INPUT,
               align="right" if 4 <= i <= 7 else None, wrap=True, unlock=True)
    st(ws.cell(r + 2, 1),
       "※ ハイブリッド / API 連携案件用（任意）。機能↔API↔実装方式↔Phase の漏れ逆引きに使う。"
       "工数は見積明細が正（この表は参考）。", F_MUTED, border=False)
    return ws


# ---------------------------------------------------------------- main
def generate(data, out_path):
    meta = data.get("meta", {})
    params = data.get("params", {})
    rates = data.get("rates", [])
    patterns = data.get("patterns", [])
    if not rates:
        raise ValueError("rates（単価表）が空です")
    if not patterns:
        raise ValueError("patterns（見積明細）が空です")

    wb = Workbook()
    wb.remove(wb.active)
    build_readme(wb, len(patterns) > 1, bool(data.get("feature_api")))
    build_settings(wb, meta, params, rates)
    arow = build_assumptions(wb, data.get("assumptions"))  # 前提条件シート（顧客提示の正本・独立）

    start = parse_date(meta.get("start_date") or meta.get("date") or
                       date.today().isoformat())
    infos = []
    for pattern in patterns:
        flat, phases = flatten_items(copy.deepcopy(pattern["items"]))
        info = build_detail(wb, pattern, phases, flat)         # 見積書（顧客に渡す明細）
        info["cells"] = build_summary(wb, pattern, info, arow)  # 見積サマリ（金額＋案件概要・連携）
        build_schedule(wb, pattern, info, rates, start)
        # 見積サマリを見積書の直前へ（隣に並べる）
        sname = info["cells"]["sheet"]
        si, di = wb.sheetnames.index(sname), wb.sheetnames.index(info["sheet"])
        if si > di:
            wb.move_sheet(sname, offset=di - si)
        infos.append(info)
    build_hearing(wb, data)   # ヒアリングシート（観点チェック・AI事前記入＋人が✓）
    build_nfr(wb, data.get("nfr"))
    # 機能×API 対応表は特殊案件（ハイブリッド/API 連携）専用 — データ指定時のみ生成
    if data.get("feature_api"):
        build_feature_api(wb, data["feature_api"])
    if len(infos) > 1:
        build_comparison(wb, infos)
        idx = wb.sheetnames.index("案比較")
        wb.move_sheet("案比較", offset=3 - idx)  # 使い方・設定・前提条件の直後へ

    wb.save(out_path)
    return out_path


def fix_fonts(xlsx_path):
    """LibreOffice で再計算保存すると、未インストールの Noto Sans JP が
    游ゴシック体等に置換されることがある。styles.xml のフォント名だけを
    ブランド正本（Noto Sans JP）へ戻す（数式・計算値・色には触れない）。"""
    import shutil
    import zipfile
    tmp = xlsx_path + ".fontfix"
    replaced = 0
    with zipfile.ZipFile(xlsx_path) as zin, \
         zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                text = data.decode("utf-8")
                for sub in ('游ゴシック体', '游ゴシック', 'Yu Gothic', 'Liberation Sans'):
                    n = text.count(f'val="{sub}"')
                    if n:
                        text = text.replace(f'val="{sub}"', f'val="{FONT}"')
                        replaced += n
                data = text.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp, xlsx_path)
    return replaced


def main():
    ap = argparse.ArgumentParser(description="bravesoft 標準見積もり Excel 生成")
    ap.add_argument("--input", help="入力 JSON ファイル")
    ap.add_argument("--out", help="出力 .xlsx パス")
    ap.add_argument("--fix-fonts", metavar="XLSX",
                    help="再計算後の .xlsx のフォント名をブランド正本（Noto Sans JP）へ復元")
    args = ap.parse_args()
    if args.fix_fonts:
        n = fix_fonts(args.fix_fonts)
        print(f"OK: {args.fix_fonts} のフォント名 {n} 箇所を {FONT} に復元")
        return
    if not (args.input and args.out):
        ap.error("--input と --out を指定（またはフォント復元は --fix-fonts）")
    with open(args.input, encoding="utf-8") as f:
        data = json.load(f)
    path = generate(data, args.out)
    print(f"OK: {path}")
    print("※ 数式の再計算を忘れずに（Excel で開くだけで OK / スクリプト運用時は recalc）。")
    print("※ recalc を LibreOffice で行った場合は --fix-fonts でフォント名を復元。")
    print("※ このファイルは社内用（社内金額・利益率入り）。顧客提出は「見積書」の社内列を折りたたんで A:I を PDF、or キットの HTML/PDF。")


if __name__ == "__main__":
    main()
